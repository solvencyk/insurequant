"""Extract Meritz Financial Group factsheet insurance KPIs to JSON.

Source: Meritz Financial Group factsheet xlsx (`MFG_202603(k).xlsx`)
Downloaded from https://m.meritzgroup.com/mo/ko/ir/ir1.do (IR011 - Quarterly results)

Scope per TODO MISC-IR-MERITZ: insurance-related KPIs (Meritz Hwajae 화재 standalone)
plus group-level RoE/BPS/EPS.

Units in source: 억원 (KRW 100M) for amounts; ratios as decimal (0.21 = 21%);
RoE/EPS as 원 (KRW). K-ICS ratio reported as multiplier (2.1255 = 212.55%).

Output: extracted_<YYYYMM>.json with structured KPI snapshot.
"""
from __future__ import annotations
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).parent
XLSX = ROOT / "factsheet_202603.xlsx"
PERIOD = "1Q26"  # latest column in source
PERIOD_KEY = "202603"


def find_period_col(ws, period: str, header_rows: range) -> int | None:
    """Locate the column index (1-based) where `period` appears in any header row."""
    for ri in header_rows:
        for ci in range(1, ws.max_column + 1):
            v = ws.cell(ri, ci).value
            if v == period:
                return ci
    return None


def find_period_col_2row(ws, year: int, quarter: str, year_row: int, q_row: int) -> int | None:
    """Find column where year_row==year AND q_row==quarter (e.g. CSM sheet 2-row header).
    Year cell may be ahead of and span over the quarter cells in the same year block."""
    cur_year: int | str | None = None
    for ci in range(1, ws.max_column + 1):
        yv = ws.cell(year_row, ci).value
        if isinstance(yv, (int, str)) and str(yv).strip().isdigit():
            cur_year = int(str(yv).strip())
        qv = ws.cell(q_row, ci).value
        if cur_year == year and qv == quarter:
            return ci
    return None


def cell(ws, row: int, col: int) -> Any:
    v = ws.cell(row, col).value
    if isinstance(v, float):
        return round(v, 6)
    return v


def extract() -> dict[str, Any]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ---- Meritz Hwajae (메리츠화재) standalone — Insurance_ALM sheet ----
    # Row layout (verified from inspection):
    # r4 = period header, r19=지급여력금액, r20=지급여력기준, r21=K-ICS 비율(구 RBC)
    # Note: source notes 1Q26 K-ICS is preliminary (잠정치)
    ws_alm = wb["Insurance_ALM"]
    col_alm = find_period_col(ws_alm, PERIOD, range(1, 6))
    assert col_alm, f"{PERIOD} not found in Insurance_ALM"
    kics_amount_eok = cell(ws_alm, 19, col_alm)         # 지급여력금액 (억원)
    kics_required_eok = cell(ws_alm, 20, col_alm)       # 지급여력기준 (억원)
    kics_ratio_mult = cell(ws_alm, 21, col_alm)         # multiplier e.g. 2.12

    # ---- CSM sheet (메리츠화재) ----
    # r10=기시 CSM, r11=신계약 CSM, r12=이자비용, r13=경험조정 등, r14=CSM 상각, r15=기말 CSM
    ws_csm = wb["CSM"]
    # CSM sheet uses 2-row header: year on r8, quarter ('1Q') on r9.
    # Each KPI block repeats this header (r20/r21, r30/r31, r40/r41).
    col_csm_mv = find_period_col_2row(ws_csm, 2026, "1Q", 8, 9)
    col_csm_nb = find_period_col_2row(ws_csm, 2026, "1Q", 20, 21)
    col_csm_mult = find_period_col_2row(ws_csm, 2026, "1Q", 30, 31)
    assert col_csm_mv and col_csm_nb and col_csm_mult, "1Q26 not found in CSM headers"
    csm_opening = cell(ws_csm, 10, col_csm_mv)
    csm_new_business = cell(ws_csm, 11, col_csm_mv)
    csm_interest_accretion = cell(ws_csm, 12, col_csm_mv)
    csm_experience_adj = cell(ws_csm, 13, col_csm_mv)
    csm_amortization = cell(ws_csm, 14, col_csm_mv)
    csm_closing = cell(ws_csm, 15, col_csm_mv)

    # New business CSM breakdown (r22-26): 보장성 인보험/물보험, 저축성, 합계
    nb_csm_protection_total = cell(ws_csm, 22, col_csm_nb)
    nb_csm_protection_personal = cell(ws_csm, 23, col_csm_nb)
    nb_csm_protection_property = cell(ws_csm, 24, col_csm_nb)
    nb_csm_savings = cell(ws_csm, 25, col_csm_nb)
    nb_csm_total = cell(ws_csm, 26, col_csm_nb)

    # NB CSM multiple (배수) r32=보장성, r35=저축성, r36=합계
    nb_csm_multiple_protection = cell(ws_csm, 32, col_csm_mult)
    nb_csm_multiple_savings = cell(ws_csm, 35, col_csm_mult)
    nb_csm_multiple_total = cell(ws_csm, 36, col_csm_mult)

    # ---- Insurance Condensed PL (메리츠화재) ----
    # First block r4-r19 is YTD (누적). For 1Q26 YTD == 1Q standalone.
    # r5=보험손익, r6=장기손익, r7=CSM상각, r8=RA, r9=예실차, r10=자동차, r11=일반,
    # r12=투자손익, r13=투자이익, r14=보험금융손익, r15=영업이익, r16=영업외, r17=세전, r18=법인세, r19=당기순이익
    ws_pl = wb["Insurance_Condensed PL"]
    col_pl = find_period_col(ws_pl, PERIOD, range(1, 6))
    assert col_pl, f"{PERIOD} not found in Insurance_Condensed PL"
    hwajae_pl = {
        "insurance_pnl_eok": cell(ws_pl, 5, col_pl),
        "longterm_pnl_eok": cell(ws_pl, 6, col_pl),
        "longterm_csm_amort_eok": cell(ws_pl, 7, col_pl),
        "longterm_ra_change_eok": cell(ws_pl, 8, col_pl),
        "longterm_experience_variance_eok": cell(ws_pl, 9, col_pl),
        "auto_pnl_eok": cell(ws_pl, 10, col_pl),
        "general_pnl_eok": cell(ws_pl, 11, col_pl),
        "investment_pnl_eok": cell(ws_pl, 12, col_pl),
        "investment_income_eok": cell(ws_pl, 13, col_pl),
        "insurance_finance_pnl_eok": cell(ws_pl, 14, col_pl),
        "operating_income_eok": cell(ws_pl, 15, col_pl),
        "nonoperating_income_eok": cell(ws_pl, 16, col_pl),
        "pretax_income_eok": cell(ws_pl, 17, col_pl),
        "tax_eok": cell(ws_pl, 18, col_pl),
        "net_income_eok": cell(ws_pl, 19, col_pl),
    }

    # ---- Insurance_Efficiency (메리츠화재) loss ratios YTD ----
    # Second block r10 header, r11=일반 손해율, r12=자동차 손해율, r13=장기 위험 손해율
    ws_eff = wb["Insurance_Efficiency"]
    col_eff = find_period_col(ws_eff, PERIOD, range(1, 11))
    assert col_eff, f"{PERIOD} not found in Insurance_Efficiency"
    loss_ratios = {
        "general_loss_ratio": cell(ws_eff, 11, col_eff),
        "auto_loss_ratio": cell(ws_eff, 12, col_eff),
        "longterm_risk_loss_ratio": cell(ws_eff, 13, col_eff),
    }

    # ---- Group highlight (메리츠금융지주 연결) ----
    # r5 header YTD; r54=ROE, r55=BPS, r56=EPS (separate header r53)
    # r21=당기순이익, r22=지배기업순이익
    ws_gh = wb["Group highlight"]
    col_gh_ytd = find_period_col(ws_gh, PERIOD, range(5, 6))
    col_gh_ratio = find_period_col(ws_gh, PERIOD, range(53, 54))
    assert col_gh_ytd, f"{PERIOD} not found in Group highlight YTD header"
    assert col_gh_ratio, f"{PERIOD} not found in Group highlight ratio header"
    group = {
        "insurance_pnl_eok": cell(ws_gh, 6, col_gh_ytd),
        "interest_pnl_eok": cell(ws_gh, 9, col_gh_ytd),
        "fee_pnl_eok": cell(ws_gh, 12, col_gh_ytd),
        "other_pnl_eok": cell(ws_gh, 15, col_gh_ytd),
        "operating_income_eok": cell(ws_gh, 19, col_gh_ytd),
        "net_income_eok": cell(ws_gh, 21, col_gh_ytd),
        "net_income_controlling_eok": cell(ws_gh, 22, col_gh_ytd),
        "roe": cell(ws_gh, 54, col_gh_ratio),
        "bps_krw": cell(ws_gh, 55, col_gh_ratio),
        "eps_krw": cell(ws_gh, 56, col_gh_ratio),
    }

    return {
        "source": {
            "ir_page": "https://m.meritzgroup.com/mo/ko/ir/ir1.do",
            "xlsx_url": "https://m.meritzgroup.com/commfiles/hld/attach/2026/20260514/202605141545243300007U.xlsx",
            "file_name": "MFG_202603(k).xlsx",
            "publisher": "Meritz Financial Group (메리츠금융지주)",
            "report_period_label": "2026년 1분기 경영실적",
            "release_date_kst": "2026-05-14",
            "encoding_note": "Source page is EUC-KR; xlsx internal strings are UTF-8.",
            "preliminary_flag": "1Q26 K-ICS ratio is preliminary (잠정치) per Insurance_ALM note",
        },
        "period": {
            "label": PERIOD,
            "key": PERIOD_KEY,  # YYYYMM end-of-period (2026.03)
        },
        "units": {
            "eok": "KRW 100,000,000 (억원)",
            "krw": "KRW (원)",
            "ratio": "decimal (e.g. 0.21 = 21%)",
            "kics_multiplier": "multiplier (e.g. 2.12 = 212%)",
        },
        "meritz_hwajae_standalone": {
            "company_kr": "메리츠화재",
            "company_en": "Meritz Fire & Marine Insurance",
            "kics": {
                "available_capital_eok": kics_amount_eok,       # 지급여력금액
                "required_capital_eok": kics_required_eok,      # 지급여력기준
                "ratio_multiplier": kics_ratio_mult,            # 지급여력비율
                "ratio_pct": round(kics_ratio_mult * 100, 2) if isinstance(kics_ratio_mult, (int, float)) else None,
                "preliminary": True,
            },
            "csm_movement_eok": {
                "opening": csm_opening,
                "new_business": csm_new_business,
                "interest_accretion": csm_interest_accretion,
                "experience_adjustment": csm_experience_adj,
                "amortization": csm_amortization,
                "closing": csm_closing,
            },
            "new_business_csm_eok": {
                "protection_total": nb_csm_protection_total,
                "protection_personal_lines": nb_csm_protection_personal,
                "protection_property_lines": nb_csm_protection_property,
                "savings": nb_csm_savings,
                "total": nb_csm_total,
            },
            "new_business_csm_multiple": {
                "protection": nb_csm_multiple_protection,
                "savings": nb_csm_multiple_savings,
                "total": nb_csm_multiple_total,
                "denominator_note": "Per IR convention: 월납월초 (initial monthly premium). See TODO decision #3.",
            },
            "pl_ytd_eok": hwajae_pl,
            "loss_ratios_ytd": loss_ratios,
        },
        "group_consolidated": {
            "entity_kr": "메리츠금융지주",
            "entity_en": "Meritz Financial Group",
            "ytd_eok": {k: v for k, v in group.items() if k not in {"roe", "bps_krw", "eps_krw"}},
            "ratios": {
                "roe": group["roe"],
                "roe_pct": round(group["roe"] * 100, 2) if isinstance(group["roe"], (int, float)) else None,
                "bps_krw": group["bps_krw"],
                "eps_krw": group["eps_krw"],
            },
            "roe_basis_note": "당기순이익은 지배기업 소유주 지분 기준; 자기자본은 기초+분기말 잔액 단순평균 (지배자본 기준)",
        },
        "internal_taxonomy_mapping": {
            "kics_ratio_pct": "meritz_hwajae_standalone.kics.ratio_pct",
            "csm_total": "meritz_hwajae_standalone.csm_movement_eok.closing",
            "nb_csm_total": "meritz_hwajae_standalone.new_business_csm_eok.total",
            "nb_csm_multiple": "meritz_hwajae_standalone.new_business_csm_multiple.total",
            "insurance_pnl": "meritz_hwajae_standalone.pl_ytd_eok.insurance_pnl_eok",
            "investment_pnl": "meritz_hwajae_standalone.pl_ytd_eok.investment_pnl_eok",
            "net_income": "meritz_hwajae_standalone.pl_ytd_eok.net_income_eok",
            "roe": "group_consolidated.ratios.roe (group only — no Hwajae-standalone RoE in factsheet)",
        },
    }


def main() -> None:
    rec = extract()
    out = ROOT / f"extracted_{PERIOD_KEY}.json"
    # Explicit utf-8 per CLAUDE.md rule 5
    out.write_text(json.dumps(rec, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {out}  ({out.stat().st_size:,} bytes)")
    # quick sanity echo
    s = rec["meritz_hwajae_standalone"]
    print(f"K-ICS {s['kics']['ratio_pct']}% (preliminary)")
    print(f"CSM closing {s['csm_movement_eok']['closing']:.1f} 억원")
    print(f"NB CSM total {s['new_business_csm_eok']['total']:.1f} 억원, multiple {s['new_business_csm_multiple']['total']:.2f}x")
    print(f"Insurance PnL {s['pl_ytd_eok']['insurance_pnl_eok']:.1f} 억원, Net income {s['pl_ytd_eok']['net_income_eok']:.1f} 억원")
    g = rec["group_consolidated"]["ratios"]
    print(f"Group RoE {g['roe_pct']}%, EPS {g['eps_krw']:.0f} 원")


if __name__ == "__main__":
    main()
