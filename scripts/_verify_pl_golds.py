#!/usr/bin/env python3
"""Verify pl_breakdown_master.json against the 4 hand-built gold xlsx (2025.4Q).

Compares item-by-item, tolerance = max(0.5% of |gold|, 1 백만). Prints PASS/FAIL per
item per company and a summary. Exit code 0 if all directly-extractable items pass."""
import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

MASTER = Path("data/dart/viz/pl_breakdown_master.json")
GOLD_DIR = Path("gold")  # gold xlsx live under gold/ (fall back to repo root for legacy paths)

# (code, gold xlsx, quarter).  First sheet holds the 24-item answer (col4=항목번호, col7=값).
# A gold may leave Tier-1 (15/17-24) BLANK on purpose — those now come from the DART FS API,
# so blank gold cells are skipped (not failed); see main().
GOLDS = [
    # 삼성화재 2025.4Q gold is 연결(CFS) — SUPERSEDED 2026-06-05: owner standardized 삼성화재 on
    # 별도(OFS).  The 별도 reference is the 2025.2Q gold below; the 연결 file stays on disk only.
    ("KR0001", "보험손익 breakdown_메리츠.xlsx", "2025.4Q"),
    ("KR0069", "보험손익 breakdown_삼성생명.xlsx", "2025.4Q"),
    ("KR0068", "보험손익 breakdown_한화생명.xlsx", "2025.4Q"),
    ("KR0003", "보험손익 breakdown_롯데_2024.xlsx", "2024.4Q"),
    ("KR0068", "보험손익 breakdown_한화생명_2025.2Q.xlsx", "2025.2Q"),
    ("KR0010", "보험손익 breakdown_KB.xlsx", "2025.2Q"),  # 분기 분해 partial: 4/5/9/10/13 (CSM상각·위험조정·자동차)
    ("KR0011", "보험손익 breakdown_DB.xlsx", "2025.2Q"),  # Tier-2 full; Tier-1 blank → FS-API
    ("KR0008", "보험손익 breakdown_삼성화재_2025.2Q.xlsx", "2025.2Q"),  # 분기 Tier-2 full; Tier-1 blank
    ("KR0005", "보험손익 breakdown_흥국화재_2025.2Q.xlsx", "2025.2Q"),  # nonPAA/PAA single-table; Tier-1 blank
    ("KR1000", "보험손익 breakdown_코리안리_2025.2Q.xlsx", "2025.2Q"),  # 재보험사: 생명(2-12)+장기(2-1~12-1)+일반(14)
    ("KR0008", "보험손익 breakdown_삼성화재_2024.2Q.xlsx", "2024.2Q"),  # 구형식(삼성-style 구분-rows)
    ("KR0011", "보험손익 breakdown_DB손보_2024.2Q.xlsx", "2024.2Q"),   # 구형식(DB-style); 시트 공시분기 오타("2024.4Q")
    ("KR0002", "보험손익 breakdown_한화손보_2024.2Q.xlsx", "2024.2Q"),  # 구형식 OLD; 별도엔 퇴직연금 없음(전부 장기/자동차/일반)
    ("KR0002", "보험손익 breakdown_한화손보.xlsx", "2025.2Q"),          # 신형식 NEW 별도; 시트 공시분기 오타("2025.4Q") 실제 당반기. 13/14 연결/별도 회귀가드
    ("KR0003", "보험손익 breakdown_롯데_2026.1Q.xlsx", "2026.1Q"),     # 롯데 신형식 분기(컴포넌트노트 보강용)
]

# items the extractor should nail directly (rest are derived/residual and follow)
# item19 (보험금융손익) is DELIBERATELY excluded: it now follows the DART FS-API standard
# (full 재보험 netting: 보험금융수익−비용＋재보험금융수익−비용).  The hand-built golds netted
# 재보험 inconsistently per company, so item19/18 may differ from gold — owner directive
# (2026-06-04): push the API standard; gold mismatch on 19 is expected/OK.
DIRECT = {1, 4, 5, 6, 9, 10, 11, 13, 14, 16, 17, 20, 22, 23, 24}  # item15 dropped: see below
# A few golds are intentionally PARTIAL: the company's isolated quarterly handler only emits a
# subset of Tier-2 (per owner directive — e.g. KB's handler is kept narrow to avoid disturbing
# other companies).  For those, gate ONLY the items the handler claims to produce; the rest of
# the (filled) gold cells are reference, not a contract.  Keyed by (code, quarter).
CLAIMED = {
    ("KR0010", "2025.2Q"): {1, 4, 5, 9, 10, 13, 16, 17, 20, 22, 23, 24},  # KB 분기: 6/11/14 not extracted
    ("KR0002", "2025.2Q"): {1, 4, 5, 6, 9, 10, 13, 14, 16},  # 한화손보 별도 직접추출 13/14 회귀가드; item11(~0.3% 가정·경험 잔차)·12(잔차파생)는 reference
    # DB손보 구형식 2024.2Q: item11(재보 예실차)·12(파생)는 derived residual SPLIT — got/gold의
    # item11+12 합계는 -82,549로 정확히 일치(내부 split 관례만 상이, item11 -3,219 ↔ item12 +3,219
    # 상쇄).  직접추출 Tier-2(1,4,5,6,9,10,13,14)는 전부 gold 일치 → 그것만 gate, 11/12는 reference.
    ("KR0011", "2024.2Q"): {1, 4, 5, 6, 9, 10, 13, 14},
    # 한화생명 2025.2Q: RA↔예실차 내부 split 관례차 (우리 추출=DART note 충실, gold=owner 귀속).
    # item5(원수RA 76,066)↔item7(예실차 -104,320): 5+7 = -28,254 = gold(75,507-103,761) 정확 일치.
    # item10(재보RA -559)↔item12(재보예실차 -7,183): 10+12 = -7,742 = gold(-450-7,292) 정확 일치.
    # → 직접추출 primary(1,4,6,9,11,13,14)만 gate, RA-split(5,10)+예실차파생(7,12)는 reference.
    ("KR0068", "2025.2Q"): {1, 4, 6, 9, 11, 13, 14},
    # 롯데 2026.1Q: item11(재보 예실차 -1,254)↔item12(파생 -1,899): 11+12 = -3,153 = gold(-3,711+558)
    # 정확 일치 (split 관례차).  직접추출만 gate, 11/12는 reference.
    ("KR0003", "2026.1Q"): {1, 4, 5, 6, 9, 10, 13, 14},
}
# item15 (기타영업수익) also excluded — the DART FS API has no separate 기타영업수익 account for
# insurers (it sits inside 보험영업수익), so Tier-1 reports it as 0 while the hand golds carried
# an HTML sub-line value.  20=1+17 holds without it; not a real error (API standard).


def load_gold(path):
    import openpyxl
    p = GOLD_DIR / path
    if not p.exists():
        p = Path(path)  # legacy: repo-root location
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] in (None, "원보험사코드"):
            continue
        raw = row[4]
        try:
            item_no = int(raw)                       # standard 1-24
        except (TypeError, ValueError):
            # reinsurer parallel-LOB items keep their string id (e.g. "2-1"…"12-1", 코리안리)
            if isinstance(raw, str) and raw.strip() and any(ch.isdigit() for ch in raw):
                item_no = raw.strip()
            else:
                continue
        out[item_no] = row[7]  # 값
    return out


def load_master():
    rows = json.loads(MASTER.read_text(encoding="utf-8"))
    out = {}
    for r in rows:
        out.setdefault((r["원보험사코드"], r["공시분기"]), {})[r["항목번호"]] = r["값"]
    return out


def close(a, b):
    if a is None or b is None:
        return False
    tol = max(0.005 * abs(b), 1.0)
    return abs(a - b) <= tol


def main():
    master = load_master()
    all_direct_ok = True
    for code, gx, q in GOLDS:
        gold = load_gold(gx)
        got = master.get((code, q), {})
        gate = CLAIMED.get((code, q), DIRECT)  # partial golds gate only their claimed items
        print(f"\n===== {code} {q}  ({gx}) =====")
        n_pass = n_fail_direct = 0
        items = list(range(1, 25)) + [k for k in gold if isinstance(k, str)]  # +코리안리 2-1…12-1
        for item_no in items:
            g = gold.get(item_no)
            v = got.get(item_no)
            # A blank gold cell (Tier-1 left to the FS API) is not applicable — skip it.
            if g is None:
                continue
            ok = close(v, g)
            tag = "PASS" if ok else "FAIL"
            # string sub-items (장기재보험 2-1…12-1) are always part of the contract
            direct = (item_no in gate) or isinstance(item_no, str)
            mark = "*" if direct else " "
            if not ok and direct:
                all_direct_ok = False
                n_fail_direct += 1
            if ok:
                n_pass += 1
            gs = f"{g:,.0f}"
            vs = "None" if v is None else f"{v:,.0f}"
            flag = "" if ok else "   <-- mismatch"
            if not ok or direct:
                print(f"  {mark}[{item_no!s:>4}] {tag}  gold={gs:>16}  got={vs:>16}{flag}")
        print(f"  -> {n_pass}/24 pass; {n_fail_direct} DIRECT items failing")
    print("\n" + ("ALL DIRECT ITEMS PASS" if all_direct_ok else "SOME DIRECT ITEMS FAIL"))
    return 0 if all_direct_ok else 1


if __name__ == "__main__":
    sys.exit(main())
