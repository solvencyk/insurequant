# -*- coding: utf-8 -*-
"""CSM waterfall gold gate: auto-discover 'CSM waterfall_*.xlsx' answer sheets in the
repo root, compare each (company, quarter) 6-stage waterfall to
data/dart/viz/csm_waterfall_master_diag.json. Gold values are 백만원; master is 억
(백만/100). Keeper gate (analog of scripts/_verify_pl_golds.py)."""
from __future__ import annotations
import glob
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
master = json.loads((ROOT / "data/dart/viz/csm_waterfall_master_diag.json").read_text(encoding="utf-8"))
cur = {(r["원보험사코드"], r["공시분기"], r["항목번호"]): r["값"] for r in master}

# name -> code (gold sheets sometimes leave 원보험사코드 blank)
NAME2CODE = {}
for r in master:
    NAME2CODE.setdefault(r["원수사명"], r["원보험사코드"])
# explicit aliases (gold sheet short names)
ALIAS = {"롯데손보": "KR0003", "DB손보": "KR0011", "농협생명": "KR0104",
         "동양생명": "KR0087", "삼성생명": "KR0069", "교보생명": "KR0073",
         "미래에셋생명": "KR0079", "푸본현대생명": "KR0083",
         "신한라이프": "KR0094", "신한라이프생명보험": "KR0094"}

LABELS = {1: "기초", 2: "신계약", 3: "이자", 4: "조정", 5: "상각", 6: "기말"}


# Some gold sheets have copy-pasted 원수사명/공시분기 cells (wrong inside the sheet).
# Override by filename: filename -> (code, quarter).
FILE_OVERRIDE = {"CSM waterfall_교보생명.xlsx": ("KR0073", "2025.4Q")}


def resolve_code(code, name):
    if code and str(code).startswith("KR"):
        return code
    return ALIAS.get(name) or NAME2CODE.get(name)


# Gold sheets split some life insurers (삼성·미래) into 사망/건강/연금저축 columns;
# the total = the SUM of those columns (handled in main). Nothing to skip.
SKIP = set()


def main():
    files = [f for f in sorted(glob.glob(str(ROOT / "CSM waterfall_*.xlsx")))
             if Path(f).name not in SKIP]
    print(f"discovered {len(files)} CSM gold sheets (company-total)\n")
    total_pass = total = 0
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb[wb.sheetnames[0]]
        gold, code, name, q = {}, None, None, None
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if row[4] is None:
                continue
            # value columns start at index 7 (single col, or 사망/건강/연금저축 split) — SUM them
            seg = [c for c in row[7:] if isinstance(c, (int, float))]
            gold[int(row[4])] = sum(seg) if seg else None
            code = code or row[0]
            name = name or row[1]
            q = q or row[6]
        ov = FILE_OVERRIDE.get(Path(f).name)
        if ov:
            code, q = ov
        else:
            code = resolve_code(code, name)
        npass = 0
        lines = []
        for i in range(1, 7):
            g = gold.get(i)
            ge = None if g is None else round(g / 100, 1)  # 백만 -> 억
            c = cur.get((code, q, i))
            ok = ge is not None and c is not None and abs(ge - c) <= max(1.0, abs(ge) * 0.01)
            npass += ok
            mark = "OK" if ok else "<<<FAIL"
            lines.append(f"    {LABELS[i]:<4} gold={str(ge):>10} parser={str(c):>10}  {mark}")
        total_pass += npass
        total += 6
        status = "PASS" if npass == 6 else f"{npass}/6"
        print(f"=== {name} {q} ({code}) -> {status} ===")
        for ln in lines:
            print(ln)
    print(f"\nTOTAL: {total_pass}/{total} stage-cells across {len(files)} gold sheets")


if __name__ == "__main__":
    main()
