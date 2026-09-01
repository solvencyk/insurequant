"""Bundle all root master tables into one reviewable .xlsx (one sheet per master).

Output: insurequant_master_tables.xlsx (repo root).
"""
from __future__ import annotations
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "insurequant_master_tables.xlsx"
FONT = "맑은 고딕"

# (json file, sheet name, description) — only real masters (diff snapshots excluded)
MASTERS = [
    ("IFRS17_BS.json", "17BS",
     "재무상태표 요약 (자산총계·부채총계·자본총계·AOCI누계액·법정준비금 3종 = 항목 1-7) long-format"),
    ("kics_disclosure.json", "K-ICS공시",
     "K-ICS 지급여력 공시 항목 (요구자본 1-35 + 시장위험 하위분해 36-46) long-format"),
    ("kics_rate_sensitivity.json", "금리민감도",
     "지급여력 금리민감도 (경과조치 적용전/후 x measure x base/±50/±100bp) — **K-ICS 쪽**. "
     "`듀레이션`·`컨벡서티`는 ±100bp 평행이동에서 유도한 유효듀레이션(년)/유효컨벡서티 "
     "(owner 2026-08-30). 양수 듀레이션 = 금리 상승 시 가치 감소. 지급여력비율 행은 두 금액의 "
     "나눗셈이라 그 미분에 듀레이션/컨벡서티라는 이름을 붙이지 않고 비워 둔다"),
    # 2026-08-30 신설. owner 상시 규칙("화면 그래프는 전부 마스터에 담는다")에 대해
    # IFRS17.html "7) 민감도(ΔCSM)" 패널만 대응 시트가 없었다. 바로 위 `금리민감도` 와
    # 이름이 비슷해 같은 것으로 오인돼 왔는데 **다른 표다** — 저쪽은 K-ICS 지급여력비율의
    # 금리 ±bp 민감도, 이쪽은 IFRS17 보험가정(사망률·장해질병·해지율·사업비) 충격의 ΔCSM.
    ("CSM_sensitivity.json", "가정민감도",
     "IFRS17 보험가정 민감도 (사망률·장해질병·해지율·사업비 충격 -> CSM변동·손익영향, 억원)"),
    ("CSM_waterfall.json", "CSM워터폴",
     "CSM 변동분석 (기초→신계약→이자부리→가정·경험조정→상각→기말)"),
    ("CSM_amortization.json", "CSM상각",
     "CSM 경과연차별 상각 스케줄"),
    ("NB_CSM_multiple.json", "신계약CSM배수",
     "신계약 CSM / 월납초회보험료 배수 (연누계)"),
    ("PL_breakdown.json", "손익분해PL",
     "손익계산서 24항목 분해 (보험·투자손익 등)"),
    ("dividend.json", "배당",
     "배당에 관한 사항 (DART alotMatter) — 항목1-7 회사단위 + 8-11 종류주(보통주/우선주)별"),
    ("kics_tier1_utilization.json", "기본자본소진율",
     "기본자본(신종자본증권) 인정한도 소진율 — SCR×15%(주력)/×10%(엄격) 한도 대비 발행잔액·인정액 "
     "(DART 사채발행현황 기준, 2026.1Q 스냅샷) long-format. 비고열=known limitation"),
    ("kics_tier2_utilization.json", "보완자본소진율",
     "보완자본(후순위채) 인정한도 소진율 — SCR×50% 한도 대비 신규발행 인정액 "
     "(DART 사채발행현황 기준, 2026.1Q 스냅샷; 2026-06-20 폐기된 구산식은 참고용 별도 항목명 행) "
     "long-format. 비고열=known limitation"),
    ("kics_capital_securities.json", "자본성증권발행현황",
     "자본성증권 한 건 단위 인정액 — 회사·구분·발행일·콜만기도래일·액면가·공시분기별 "
     "기본자본인정액/보완자본인정액. 종전에는 회사 단위 집계값만 있어 어느 증권이 얼마를 "
     "인정받는지 되짚을 수 없었다(owner 2026-09-01 설계). 신종은 SCR×15% 한도를 "
     "경과조치분·신규분이 공유하며 발행일 순으로 채우고 초과분은 보완자본으로 분류, "
     "후순위는 잔존만기 5년 미만부터 매년 20%p 계단식 차감"),
    ("kics_forward_capital.json", "자본비율전망",
     "자본비율 5년 전망(2026~2030 연도말) — 콜옵션 도래 채무성자본 차감 + SCR 선형보간 시뮬레이션 "
     "(baseline 2026.1Q, status!=ok 회사 제외) long-format, 공시분기 칸에 전망연도. 비고열=known limitation"),
]

NUMERIC_COLS = {"값", "-100bp", "-50bp", "base", "+50bp", "+100bp",
                "액면가_억", "잔액_억", "기본자본인정액_억", "보완자본인정액_억",
                "기본자본한도_억", "SCR_억", "보완자본인정율",
                "상각액", "신계약CSM_연누계", "월납월초보험료_연누계", "신계약CSM배수_연누계",
                "CSM변동", "당기손익영향", "자본영향",
                "듀레이션", "컨벡서티"}
TEXT_COLS = {"원보험사코드", "원수사명", "티커", "생손보여부", "공시분기",
             "항목명", "경과조치여부", "measure구분", "경과차년", "종류주", "섹션", "레벨",
             "기준일", "순번", "위험구분", "충격수준", "비고"}


def coerce(df):
    for c in df.columns:
        if c in NUMERIC_COLS:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        elif c == "항목번호":
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
        else:
            df[c] = df[c].astype("string")
    return df


# ---------------------------------------------------------------------------
# Flatteners — kics_tier1_utilization.json / kics_tier2_utilization.json /
# kics_forward_capital.json are NOT already long-format (unlike the 8 masters
# above, which are flat lists-of-dicts straight off disk). Each is a per-company
# snapshot (tier1/2) or per-company x per-year-projection nest (forward). These
# reshape them into the SAME (원보험사코드,원수사명,티커,생손보여부,공시분기,항목명,값)
# shape as everything else, so xlsx/sync/coerce all stay generic.
#
# inbox/parser/20260829T0100Z (orchestrator, owner-approved schema 2026-08-29):
#   tier1/tier2: 공시분기 = snapshot quarter ("2026.1Q"), 항목명 splits the metric.
#   forward:     공시분기 = projection year ("2026".."2030") — plain 4-digit strings,
#                never mixed with "YYYY.NQ" quarter labels in the same column (checked
#                against scripts/export_public_sheets.py:41-44's non-standard-label
#                sort concern; see this ticket's report for why it doesn't apply here).
# One added column, "비고" (free-text, blank for most rows): the ticket requires the
# known extraction limitations (tier1 numerator partly BS-fallback, forward call-date
# partly estimated, low-confidence rows, deprecated tier2 proxy metric) to travel WITH
# the numbers if someone copies just the numeric rows out — a header-only note would not
# survive that. Every other column name is reused verbatim from the existing 8 sheets.
# ---------------------------------------------------------------------------

def _kics_company_lookup() -> dict:
    """원보험사코드 -> (원수사명, 티커, 생손보여부), sourced from kics_disclosure.json.

    Every insurer in the tier1/tier2/forward-capital cohorts already has K-ICS
    disclosure rows (verified 2026-08-29: all 39 tier1/tier2 codes and all 38
    forward_capital codes are a subset of kics_disclosure.json's 원보험사코드 set) —
    this reuses that already-verified truth instead of standing up a new registry.
    """
    data = json.loads((REPO / "kics_disclosure.json").read_text(encoding="utf-8"))
    lut: dict = {}
    for r in data:
        code = r.get("원보험사코드")
        if code and code not in lut:
            lut[code] = (r.get("원수사명"), r.get("티커"), r.get("생손보여부"))
    return lut


def _company_base(code, fallback_name, lut: dict) -> dict:
    name, ticker, biztype = lut.get(code, (None, None, None))
    return {"원보험사코드": code, "원수사명": name or fallback_name,
            "티커": ticker, "생손보여부": biztype}


def _call_fallback_codes() -> set:
    """원보험사코드 set with >=1 OUTSTANDING bond whose call_source != 'disclosed'
    (call option not found in the filing -- forward_capital_simulation.py L150 derives/
    estimates effective_call_date instead, e.g. issue+5y or rating-history heuristics).
    Computed live off data/bonds/capital_securities_fy2025.json, not hardcoded — census
    verified 2026-08-29: 20 of 39 companies (see this ticket's report for the full list).
    """
    path = REPO / "data" / "bonds" / "capital_securities_fy2025.json"
    if not path.exists():
        return set()
    doc = json.loads(path.read_text(encoding="utf-8"))
    out = set()
    for c in doc.get("companies", []):
        for b in c.get("bonds", []):
            if b.get("outstanding_mn") and b.get("call_source") != "disclosed":
                out.add(c.get("code"))
    return out


_TIER1_ITEMS = [
    ("scr_eok", "지급여력기준금액(SCR)"),
    ("tier1_hybrid_limit_eok", "기본자본 인정한도(SCR x 15%)"),
    ("tier1_hybrid_limit_strict_eok", "기본자본 인정한도(엄격, SCR x 10%)"),
    ("tier1_hybrid_issued_eok", "신종자본증권 발행잔액"),
    ("tier1_hybrid_recognized_eok", "신종자본증권 인정액"),
    ("tier1_hybrid_overflow_eok", "신종자본증권 한도초과분"),
    ("legacy_hybrid_transition_eok", "경과조치 신종자본증권(구산정)"),
    ("tier1_grandfathered_hybrid_eok", "경과조치 인정 신종자본증권"),
    ("utilization_pct", "기본자본 소진율"),
    ("utilization_pct_strict", "기본자본 소진율(엄격)"),
]
# tier1_hybrid_excess_eok dropped: null for 39/39 rows (verified 2026-08-29) — an
# always-empty column carries no information as a row.
# utilization_pct_raw dropped: byte-identical to utilization_pct for 39/39 rows
# (verified 2026-08-29) — pure duplicate.

_TIER1_ISSUED_NOTE = ("신종자본증권 발행잔액이 DART 채권별 공시에서 직접 추출되지 않아 "
                       "K-ICS 경과조치 인정액(BS)으로 대체 — 실제 미상환 발행잔액과 다를 수 있음")
_TIER1_NOTE_FIELDS = {"tier1_hybrid_issued_eok", "tier1_hybrid_recognized_eok", "utilization_pct"}

# 2026-08-29 코디네이터 후속 요청: 100% 초과 13행(실측: primary>100 6개사 + strict>100
# 7개사, 6개사는 두 항목 다 초과라 중복) 비고가 비어 있어 "엑셀만 받아본 사람은 파싱오류로
# 읽는다"는 지적 + 소진율/소진율(엄격) 차이가 시트만 봐서 안 읽힌다는 지적. 근거는 owner
# 승인 정의서 docs/tier1_hybrid_utilization_definition.md("왜 캡이 없나", "소진율 필드
# 의미" 절, 2026-06-14/2026-08-25 owner 결정) — 추측 아니고 그 문서를 그대로 인용.
_TIER1_BASIS_NOTE = {
    "utilization_pct": (
        "인정한도 기준=SCR×15%(신종자본증권이 조건부자본증권일 때 상향되는 한도, KIRI "
        "2024-14). 옆 '기본자본 소진율(엄격)'은 같은 분자를 SCR×10%(비조건부 원칙한도)로 "
        "나눈 값이라 정의상 이 값의 1.5배로 더 높게 나온다."
    ),
    "utilization_pct_strict": (
        "인정한도 기준=SCR×10%(신종자본증권의 비조건부 원칙한도, KIRI 2024-14 p12). "
        "'기본자본 소진율'(SCR×15%, 조건부자본증권 인정 시 상향한도)의 1.5배 — 분자는 "
        "동일, 분모만 더 작다."
    ),
}
_TIER1_OVER100_NOTE = (
    "100%초과는 파싱오류 아님 — 분자(발행 인정액, DART 채권별)와 분모(인정한도, K-ICS "
    "공시 SCR 기반)가 서로 독립 소스라 설계상 100%로 묶이지 않는다(캡 없음, owner "
    "2026-06-14 결정 — docs/tier1_hybrid_utilization_definition.md). 화면(K-ICS.html "
    "자본증권 도넛)에는 '100%+'로 표기됨."
)


def _flatten_tier1(raw: dict) -> list:
    lut = _kics_company_lookup()
    q = raw.get("quarter")
    rows = []
    for r in raw.get("results", []):
        base = _company_base(r.get("code"), r.get("company"), lut)
        issued_note = _TIER1_ISSUED_NOTE if r.get("issued_source") == "missing" else ""
        for field, label in _TIER1_ITEMS:
            row = dict(base)
            row["공시분기"] = q
            row["항목명"] = label
            v = r.get(field)
            row["값"] = v
            notes = []
            if field in _TIER1_NOTE_FIELDS and issued_note:
                notes.append(issued_note)
            basis = _TIER1_BASIS_NOTE.get(field)
            if basis:
                notes.append(basis)
                if (v or 0) > 100.0:
                    notes.append(_TIER1_OVER100_NOTE)
            row["비고"] = " / ".join(notes)
            rows.append(row)
    return rows


_TIER2_ITEMS = [
    ("tier2_limit_eok", "보완자본 인정한도(SCR x 50%)"),
    ("numerator_eok", "보완자본 인정한도 소진 분자"),
    ("utilization_pct", "보완자본 소진율"),
    ("tier2_eok", "보완자본 총액(경과조치 후)"),
    ("pre_limit_eok", "보완자본 한도적용전 잔액"),
    ("lapse_excess_eok", "해약환급금준비금 상당액 초과분(한도제외)"),
    ("hybrid_eok", "기발행 신종자본증권(보완자본 재분류분)"),
    ("subordinated_eok", "기발행 후순위채무"),
    ("new_subordinated_gross_eok", "신규(2023~) 후순위채 발행총액"),
    ("new_subordinated_recognized_eok", "신규(2023~) 후순위채 인정액"),
    ("tier1_overflow_into_tier2_eok", "기본자본 한도초과분의 보완자본 재분류"),
    ("grandfathered_hybrid_eok", "경과조치 인정 신종자본증권(레거시)"),
    ("grandfathered_subordinated_eok", "경과조치 인정 후순위채무(레거시)"),
    ("proxy_utilization_pct", "보완자본 소진율(구 산식, 참고용)"),
]
# 2026-08-29 census (validate_data_contract.py CHECK4 domain-identity, live): RED=0
# YELLOW=0 on tier1/tier2 today -- the "4개사 분자 불안정" issue this ticket's brief
# describes matches output/tier2_utilization/outlier_report_20261Q.json (2026-06-16,
# STALE — 5 companies over the OLD proxy_utilization_pct metric: 동양240%/하나손해235%/
# KB218%/악사197%/미래126%), already fixed by inbox/_resolved/20260620T0238Z (numerator
# switched from item3-proxy to DART bond-level issuance). Current utilization_pct has 0
# companies outside [0,100] (39/39 in range). proxy_utilization_pct is kept in the JSON
# on purpose as an audit trail of the old (broken) methodology, not a live defect — the
# note below documents that distinction so nobody downstream mistakes it for current.
_TIER2_PROXY_NOTE = ("2026-06-20 폐기된 구 산식(item3 총보완자본 raw 기준) — 한도제외 대상인 "
                      "해약환급금준비금 초과분이 섞여 부풀었던 값(예: 동양생명 240%, KB손해보험 218%). "
                      "신종/후순위 발행잔액이 아니다. 현재 헤드라인은 '보완자본 소진율' 행(분자=DART "
                      "발행잔액 기반, 2026-08-29 census: 39개사 전부 0~100% 안).")


def _flatten_tier2(raw: dict) -> list:
    lut = _kics_company_lookup()
    q = raw.get("quarter")
    rows = []
    for r in raw.get("results", []):
        base = _company_base(r.get("code"), r.get("company"), lut)
        for field, label in _TIER2_ITEMS:
            row = dict(base)
            row["공시분기"] = q
            row["항목명"] = label
            row["값"] = r.get(field)
            row["비고"] = _TIER2_PROXY_NOTE if field == "proxy_utilization_pct" else ""
            rows.append(row)
    return rows


_FORWARD_ITEMS = [
    ("ratio_pct", "지급여력비율 전망"),
    ("basic_ratio_pct", "기본자본비율 전망"),
    ("capital_eok", "가용자본 전망"),
    ("basic_capital_eok", "기본자본 전망"),
    ("scr_eok", "지급여력기준금액(SCR) 전망"),
    ("cumulative_bond_dedu_eok", "누적 채무성자본 콜상환 차감액"),
    ("cumulative_tier1_dedu_eok", "누적 기본자본 신종 콜상환 차감액"),
    ("hybrid_remaining_eok", "잔존 신종자본증권(기본자본 인정분)"),
    ("hybrid_tier1_eok", "기본자본 인정 신종자본증권(연도말)"),
    ("hybrid_tier2_overflow_eok", "기본자본 한도초과 보완자본 이관분(연도말)"),
    ("hybrid_limit_eok", "기본자본 신종 인정한도(연도별 SCR 기준)"),
]
# baseline/baseline_2025_4Q intentionally NOT emitted as rows: they duplicate items
# 1/14/27 already in the K-ICS공시 sheet at 공시분기=2026.1Q — re-flattening them here
# would (a) add no new information and (b) force mixing "2026.1Q"-style and plain-year
# "2026" labels in the same 공시분기 column for one company, which is exactly the
# non-standard-label sort risk export_public_sheets.py:41-44 documents (checked; see
# this ticket's report). scr_interp_progress dropped: pure (year-2025)/7 restatement of
# the year already carried by 공시분기. capacity_exhausted/basic_capacity_exhausted
# folded into 비고 (see below) rather than emitted as 0/1 rows.


def _flatten_forward_capital(raw: list) -> list:
    lut = _kics_company_lookup()
    call_fallback = _call_fallback_codes()
    rows = []
    for r in raw:
        if r.get("status") != "ok":
            continue
        code = r.get("insurer_code")
        base = _company_base(code, r.get("insurer_name"), lut)
        conf = r.get("confidence") or {}
        notes = []
        if r.get("bond_coverage") == "no_bonds_in_dart":
            notes.append("DART 자본성증권 발행 원자료 없음 — 콜상환 차감 없이 SCR 선형보간만 반영")
        if code in call_fallback:
            notes.append("콜옵션 미공시 채권 포함 — 콜일자를 발행일+5년/신용등급이력 등으로 추정 "
                         "(원문 미기재), 실제 콜상환 시점과 다를 수 있음")
        if conf.get("level") == "low":
            reasons = "; ".join(conf.get("reasons") or [])
            notes.append(f"신뢰도 낮음(발행잔액 vs BS 괴리) — {reasons}" if reasons else "신뢰도 낮음")
        base_note = " / ".join(notes)
        for p in r.get("projections", []):
            for field, label in _FORWARD_ITEMS:
                row = dict(base)
                row["공시분기"] = str(p.get("year"))
                row["항목명"] = label
                row["값"] = p.get(field)
                note = base_note
                if field == "ratio_pct" and p.get("capacity_exhausted"):
                    note = (note + " / " if note else "") + "자본잠식(가용자본<=0) — 비율 0%로 캡 표시"
                if field == "basic_ratio_pct" and p.get("basic_capacity_exhausted"):
                    note = (note + " / " if note else "") + "기본자본 잠식(가용 기본자본<=0) — 비율 0%로 캡 표시"
                row["비고"] = note
                rows.append(row)
    return rows


def _flatten_capital_securities(doc):
    """이미 한 건 = 한 행이라 rows 만 꺼내면 된다(집계 해체가 필요 없다)."""
    rows = doc.get("rows", []) if isinstance(doc, dict) else (doc or [])
    out = []
    for r in rows:
        row = {k: v for k, v in r.items() if k != "flags_missing"}
        miss = r.get("flags_missing") or []
        row["비고"] = ("원천에 " + "·".join(miss) + " 플래그가 없어 신규분 10%/15% 판정과 "
                       "상환촉진 유인 콜 판정은 미확정") if miss else ""
        out.append(row)
    return out


FLATTEN = {
    "kics_capital_securities.json": _flatten_capital_securities,
    "kics_tier1_utilization.json": _flatten_tier1,
    "kics_tier2_utilization.json": _flatten_tier2,
    "kics_forward_capital.json": _flatten_forward_capital,
}


def main():
    frames = []
    for fn, sheet, desc in MASTERS:
        data = json.loads((REPO / fn).read_text(encoding="utf-8"))
        if fn in FLATTEN:
            data = FLATTEN[fn](data)
        df = pd.DataFrame(data)
        df = coerce(df)
        frames.append((sheet, df, fn, desc))

    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        # placeholder index first (filled after we know counts)
        pd.DataFrame({"_": []}).to_excel(xw, sheet_name="요약", index=False)
        for sheet, df, _fn, _desc in frames:
            df.to_excel(xw, sheet_name=sheet, index=False)

    wb = load_workbook(OUT)

    # ---- 요약 (index) sheet ----
    idx = wb["요약"]
    idx.delete_cols(1, 4)
    idx["A1"] = "Insurequant 마스터테이블 통합"
    idx["A1"].font = Font(name=FONT, bold=True, size=14)
    headers = ["시트", "마스터 파일", "행수", "설명"]
    idx.append([])
    idx.append(headers)
    hdr_row = 3
    for sheet, df, fn, desc in frames:
        idx.append([sheet, fn, len(df), desc])
    idx.append(["합계", "", sum(len(df) for _s, df, _f, _d in frames), ""])
    # style index
    for c in range(1, 5):
        cell = idx.cell(row=hdr_row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    total_row = hdr_row + 1 + len(frames)
    for r in range(hdr_row + 1, total_row + 1):
        bold = (r == total_row)
        for c in range(1, 5):
            idx.cell(row=r, column=c).font = Font(name=FONT, bold=bold)
        idx.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        idx.cell(row=r, column=3).number_format = "#,##0"
    idx.column_dimensions["A"].width = 16
    idx.column_dimensions["B"].width = 30
    idx.column_dimensions["C"].width = 10
    idx.column_dimensions["D"].width = 70
    idx.sheet_view.showGridLines = False

    # ---- data sheets ----
    thin = Side(style="thin", color="D9D9D9")
    for sheet, df, _fn, _desc in frames:
        ws = wb[sheet]
        ncol = ws.max_column
        nrow = ws.max_row
        # header style
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name=FONT, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        # body font + number format
        cols = [ws.cell(row=1, column=c).value for c in range(1, ncol + 1)]
        for c, name in enumerate(cols, start=1):
            letter = ws.cell(row=1, column=c).column_letter
            for r in range(2, nrow + 1):
                cl = ws.cell(row=r, column=c)
                cl.font = Font(name=FONT)
                if name in NUMERIC_COLS and isinstance(cl.value, (int, float)):
                    cl.number_format = "#,##0.##;(#,##0.##);-"
            # column width (approx from header + name length)
            base_w = {"원수사명": 22, "항목명": 30, "설명": 60}.get(name, 0)
            ws.column_dimensions[letter].width = base_w or max(11, min(28, len(str(name)) + 6))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ncol).column_letter}{nrow}"

    # order: 요약 first
    wb.move_sheet("요약", -(wb.sheetnames.index("요약")))
    wb.save(OUT)
    print(f"wrote {OUT.name}: {len(frames)} master sheets + 요약")
    for sheet, df, _fn, _desc in frames:
        print(f"  {sheet:14s} {len(df):6d} rows x {len(df.columns)} cols")


if __name__ == "__main__":
    main()
