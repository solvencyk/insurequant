# -*- coding: utf-8 -*-
"""Parse Hanwha Life (KR0068) quarterly factsheets -> 신계약 CSM 배수 series.

Each factsheet has a sheet '신계약 CSM 및 수익성' with three blocks:
  ■ 신계약 CSM                          (단위: 십억원)  -> amount
  ■ APE 대비 신계약 CSM 배수                              -> (not used)
  ■ 월평균 월초 대비 신계약 CSM 배수      (신계약CSM ÷ 월초) -> the task metric

Within each block: a period sub-header row (e.g. '2023.1~3' .. '2025.10~12')
maps each calendar quarter to a column index; data rows (전체/보장성/종신/
건강|일반보장/연금/저축) carry the values at those columns.

Newer factsheets carry comparative columns back ~5 quarters; we union all
factsheets, preferring the *latest* factsheet's (possibly restated) value for
each quarter, and record which source files reported it.

Output: data/ir/series/KR0068_한화생명.json
"""
import json
import re
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402
import xlrd  # noqa: E402

STAGE = Path("data/ir/decks/hanwha_life").resolve()
OUT = Path("data/ir/series/KR0068_한화생명.json").resolve()
CSM_SHEET = "신계약 CSM 및 수익성"

# requested range only
WANT_YEARS = range(2023, 2027)

PERIOD_RE = re.compile(r"(\d{4})\.(\d{1,2})~(\d{1,2})")


def period_to_q(label):
    """'2023.1~3' -> (2023,'1Q'); only accept exact quarter ranges."""
    m = PERIOD_RE.search(str(label))
    if not m:
        return None
    y, a, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
    qmap = {(1, 3): "1Q", (4, 6): "2Q", (7, 9): "3Q", (10, 12): "4Q"}
    q = qmap.get((a, b))
    if not q:
        return None
    return y, q


def read_grid(path):
    """Return list-of-rows (each a list of cell values) for the CSM sheet,
    handling both .xlsx (openpyxl) and legacy .xls (xlrd)."""
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        name = next((s for s in wb.sheetnames if s.strip() == CSM_SHEET), None)
        if name is None:
            return None
        ws = wb[name]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        wb = xlrd.open_workbook(path)
        name = next((s for s in wb.sheet_names() if s.strip() == CSM_SHEET), None)
        if name is None:
            return None
        ws = wb.sheet_by_name(name)
        return [ws.row_values(ri) for ri in range(ws.nrows)]


def cell_str(c):
    return "" if c is None else str(c).strip()


def parse_factsheet(path):
    """-> dict: {(year,'NQ'): {'nb_csm_eok':.., 'multiple_total':..,
                               'multiple_by_product':{...}}}"""
    grid = read_grid(path)
    if grid is None:
        return {}
    # locate the three block headers and their period-header row
    blocks = {}  # block_name -> (header_row_idx)
    for ri, row in enumerate(grid):
        first = cell_str(row[0]) if row else ""
        if first.startswith("■ 신계약 CSM"):
            blocks["amount"] = ri
        elif "월평균 월초 대비 신계약 CSM 배수" in first:
            blocks["mult"] = ri

    def col_map(header_row_idx):
        """Find the period sub-header row within ~4 rows below the block header,
        return {col_idx: (year,'NQ')}."""
        for off in range(1, 6):
            ri = header_row_idx + off
            if ri >= len(grid):
                break
            row = grid[ri]
            cm = {}
            for ci, c in enumerate(row):
                pq = period_to_q(cell_str(c))
                if pq:
                    cm[ci] = pq
            if cm:
                return ri, cm
        return None, {}

    def data_rows_after(period_row_idx, n=8):
        """Return data rows (the labelled metric rows) after the period header."""
        out = []
        for ri in range(period_row_idx + 1, min(period_row_idx + 1 + n, len(grid))):
            row = grid[ri]
            label = cell_str(row[0]) if row else ""
            # stop at next block / footnote
            if label.startswith("■") or label.startswith("※"):
                break
            # label may be in col 0 (전체/신계약 CSM) or col 1 (보장성 등)
            lab = label or (cell_str(row[1]) if len(row) > 1 else "")
            out.append((lab.strip(), row))
        return out

    def to_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    result = {}

    # --- amount block (십억원 -> 억원 x10) ---
    if "amount" in blocks:
        prow, cmap = col_map(blocks["amount"])
        for lab, row in data_rows_after(prow):
            if lab.replace(" ", "") != "신계약CSM":
                continue
            for ci, (y, q) in cmap.items():
                if y not in WANT_YEARS:
                    continue
                # skip the 2022.4Q-as-annual anomaly handled by WANT_YEARS>=2023
                val = to_float(row[ci]) if ci < len(row) else None
                if val is None:
                    continue
                result.setdefault((y, q), {})["nb_csm_eok"] = round(val * 10, 2)
            break

    # --- month-avg multiple block ---
    if "mult" in blocks:
        prow, cmap = col_map(blocks["mult"])
        prod_label = {
            "보장성": "보장성",
            "종신": "종신",
            "일반보장": "건강",  # renamed to 건강 from FY25 Q2
            "건강": "건강",
            "연금/저축": "연금저축",
            "연금": "연금저축",
        }
        for lab, row in data_rows_after(prow):
            clean = lab.replace(" ", "")
            is_total = ("전체수익성" in clean) or clean == "전체"
            for ci, (y, q) in cmap.items():
                if y not in WANT_YEARS:
                    continue
                val = to_float(row[ci]) if ci < len(row) else None
                if val is None:
                    continue
                slot = result.setdefault((y, q), {})
                if is_total:
                    slot["multiple_total"] = round(val, 3)
                else:
                    key = None
                    for k, v in prod_label.items():
                        if clean.startswith(k):
                            key = v
                            break
                    if key:
                        slot.setdefault("multiple_by_product", {})[key] = round(val, 3)
    return result


def main():
    files = sorted(p for p in STAGE.glob("*factsheet*")
                   if p.suffix.lower() in (".xls", ".xlsx"))
    # sort oldest->newest by FY token so newest overwrites (restated wins)
    def fy_key(p):
        m = re.search(r"FY(\d{4})_Q(\d)", p.name)
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)
    files.sort(key=fy_key)

    merged = {}  # (y,q) -> record (with source_files set)
    for f in files:
        try:
            parsed = parse_factsheet(f)
        except Exception as e:
            print(f"  ERR {f.name}: {type(e).__name__}: {e}")
            continue
        for (y, q), rec in parsed.items():
            key = f"{y}.{q}"
            slot = merged.setdefault(key, {"source_files": set()})
            # latest file wins for numeric values
            for k, v in rec.items():
                slot[k] = v
            slot["source_files"].add(f.name)
        if parsed:
            print(f"  {f.name}: quarters {sorted(f'{y}.{q}' for (y,q) in parsed)}")

    # restrict to requested range FY2023.1Q .. FY2026.1Q
    def in_range(key):
        y, q = key.split(".")
        y = int(y); qn = int(q[0])
        if y < 2023 or y > 2026:
            return False
        if y == 2026 and qn > 1:
            return False
        return True

    series = {}
    for key in sorted(k for k in merged if in_range(k)):
        rec = merged[key]
        src = sorted(rec.pop("source_files"))
        out = {
            "multiple_total": rec.get("multiple_total"),
            "multiple_by_product": rec.get("multiple_by_product", {}),
            "nb_csm_eok": rec.get("nb_csm_eok"),
            "source_file": src[-1] if src else None,
            "source_files": src,
        }
        series[key] = out

    doc = {
        "company": "한화생명",
        "kr": "KR0068",
        "sector": "Life",
        "metric": "월평균 월초 대비 신계약 CSM 배수 (배) = 신계약 CSM ÷ 월평균 월초 — "
                  "Hanwha Life IR factsheet sheet '신계약 CSM 및 수익성' (별도)",
        "units": {"nb_csm_eok": "KRW 100M (억원)", "multiple": "배 (x), 월초 기준"},
        "product_note": "multiple_by_product: 보장성(=종신+건강), 종신, 건강(=옛 일반보장), 연금저축",
        "series": series,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nwrote {OUT}")
    for k in sorted(series):
        s = series[k]
        print(f"  {k}: total={s['multiple_total']}  보장성={s['multiple_by_product'].get('보장성')}  "
              f"nb_csm={s['nb_csm_eok']}억")
    missing = []
    for y in range(2023, 2027):
        for q in range(1, 5):
            if y == 2026 and q > 1:
                break
            if f"{y}.{q}Q" not in series:
                missing.append(f"{y}.{q}Q")
    print("missing:", missing or "none")


if __name__ == "__main__":
    main()
