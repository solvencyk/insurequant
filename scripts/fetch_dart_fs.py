#!/usr/bin/env python3
"""DART standardized financial-statement API → PL Tier-1 (income statement) source.

Replaces the fragile HTML income-statement parsing.  DART's fnlttSinglAcntAll.json
returns the 포괄손익계산서 keyed by STANDARD account_id (ifrs-full_* / dart_*), so the
mapping is robust even though account_nm varies per insurer (보험손익 vs 보험서비스결과,
당기순이익 vs 반기순이익 vs 연결반기순이익).  Verified against the hand-built golds
(한화생명·KB·삼성·메리츠) — exact match via thstrm_add_amount(누적) for 반기/분기.

Owner directive (2026-06-04): use the FS API for Tier-1 fleet-wide.  Tier-2 (the IFRS17
decomposition: CSM상각/위험조정/예실차, 장기/자동차/일반) is footnote-only → stays hand-parsed.

corp_code is resolved by NAME at runtime (CORPCODE.xml), per the no-permanent-map rule.
Raw API JSON is cached under data/dart/_fs_api_cache/ (external network data)."""
from __future__ import annotations
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path.cwd()))
from src.ifrs17.opendart_client import OpenDARTClient  # noqa: E402

CACHE = Path("data/dart/_fs_api_cache")
REPRT = {"1Q": "11013", "2Q": "11012", "3Q": "11014", "4Q": "11011"}
# Basis (fs_div): owner directive 2026-08 — CSM/PL masters are unified on 별도(OFS).
# Default OFS; CFS is a same-quarter FALLBACK only (when OFS has no income statement at
# all), never a preferred primary — BASIS_CFS stays empty.
# NOTE: 삼성화재(KR0008) 연결→별도 on 2026-06-05 (owner: 별도 답지 2025.2Q; 연결은 해외 일반/자동차
# 자회사를 끌어와 LOB 분해를 왜곡 — 별도가 본체 보험손익 분해에 맞음).
# 2026-08-26 (inbox/parser/20260825T1415Z follow-up): 삼성생명·메리츠 REMOVED from this set.
# Both were audited against raw XBRL ACONTEXT tags (…ConsolidatedAndSeparateFinancial
# StatementsAxis_ifrs-full_{Consolidated,Separate}Member) — the master's item24 exactly
# matched the ConsolidatedMember cell for both (삼성생명 2025.4Q: 2,451,515 vs SeparateMember
# 1,699,762 candidate value fetched fresh below; 메리츠 2025.4Q: 1,692,866.810136 vs
# SeparateMember 1,681,024.330229), i.e. the "gold=연결" comment was the same class of bug
# as the CSM line-65535 block-selection defect (b2293c8) — a stale pre-owner-directive
# assumption, not a verified exception.  OFS income-statement data exists and parses cleanly
# for both codes across FY2023–FY2025 (spot-checked), so this is not a coverage regression.
BASIS_CFS = set()
# name-search aliases (Korean transliteration the substring search can't reach) — NOT a
# permanent KR↔corp map; just better search terms.
ALIAS = {"KB라이프생명": "케이비라이프생명보험", "IBK연금보험": "아이비케이연금보험"}

# Tier-1 item → exact standard account_id (포괄손익계산서, sj_div IS/CIS)
ACCT = {
    1: "ifrs-full_InsuranceServiceResult",                       # 보험손익/보험서비스결과
    17: "dart_InvestmentIncomeExpenses",                         # 투자손익
    20: "ifrs-full_ProfitLossFromOperatingActivities",           # 영업이익
    22: "ifrs-full_ProfitLossBeforeTax",                         # 세전이익
    23: "ifrs-full_IncomeTaxExpenseContinuingOperations",        # 법인세
    24: "ifrs-full_ProfitLoss",                                  # 당기순/반기순/분기순
    16: "dart_OtherOperatingExpenseInsurance",                   # 기타사업비용
}
# item19 (보험금융손익) = 보험금융수익 − 비용 + 재보험금융수익 − 비용.  Some insurers have no
# 보험금융수익 account (한화생명) → treat missing as 0 (per owner note).
FIN = {
    "ins_inc": "dart_InsuranceFinanceIncomeFromInsuranceContractsIssuedRecognisedInProfitOrLoss",
    "ins_exp": "dart_InsuranceFinanceExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss",
    "re_inc": "dart_FinanceIncomeFromReinsuranceContractsHeldRecognisedInProfitOrLoss",
    "re_exp": "dart_FinanceExpensesFromReinsuranceContractsHeldRecognisedInProfitOrLoss",
}
NONOP_INC, NONOP_EXP = "dart_NonOperatingIncome", "dart_NonOperatingExpense"
# hidden keys for Tier-2 item3/8 derivation (보험수익/비용 grand totals), startswith-matched
IS_PREFIX = {
    "_is_rev": "ifrs-full_InsuranceRevenue",
    "_is_cost": "ifrs-full_InsuranceServiceExpenses",
    "_is_rerev": "ifrs-full_IncomeFromAmountsRecovered",
    "_is_recost": "ifrs-full_ExpensesFromAllocationOfPremiums",
}

# Items 25-31 (총포괄손익 extension, owner ticket inbox/parser/20260828T0113Z) — same sj_div
# CIS rows already scanned into `vals` above, keyed by account_id (NOT account_nm: 삼성생명
# alone spells item25 THREE different ways across FY2023-2026 -- "기타포괄손익" /
# "법인세비용차감후기타포괄손익" / … -- while account_id stays ifrs-full_OtherComprehensiveIncome
# throughout; an exact-nm-match census silently drops the whole company, see ticket §작업1).
# Mapping chosen from a full-universe 356-cell census (scripts/_probes/census_oci_labels_pass{1,2}.py,
# artifacts/parser/oci_label_census_pass{1,2}.json) and verified via the item24+25=31 identity,
# which closes to EXACT 0.000 residual across all 282 CIS-bearing cells -- strong confirmation
# these are the right tags, not a same-shape decoy.
ACCT_OCI = {
    25: "ifrs-full_OtherComprehensiveIncome",                     # 기타포괄손익
    26: "ifrs-full_OtherComprehensiveIncomeNetOfTaxFinancialAssetsMeasuredAtFairValueThroughOtherComprehensiveIncome",  # FVOCI 채무증권 평가손익
    27: "ifrs-full_OtherComprehensiveIncomeNetOfTaxInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLossThatWillBeReclassifiedToProfitOrLoss",  # 보험계약금융손익(OCI)
    28: "ifrs-full_OtherComprehensiveIncomeNetOfTaxCashFlowHedges",  # 위험회피 파생상품 평가손익
    29: "ifrs-full_OtherComprehensiveIncomeNetOfTaxGainsLossesFromInvestmentsInEquityInstruments",  # FVOCI 지분증권 평가손익
    30: "ifrs-full_OtherComprehensiveIncomeNetOfTaxFinanceIncomeExpensesFromReinsuranceContractsHeldExcludedFromProfitOrLoss",  # 재보험금융손익(OCI)
    31: "ifrs-full_ComprehensiveIncome",                           # 총포괄손익
}
# item28 fallback ids: 교보생명(KR0073) drops the standard CashFlowHedges tag from FY2025.1Q
# onward and reuses a bare Gain/Loss-derivatives concept AS the signed net line (raw-confirmed:
# dart_GainFromDerivativesHeldForHedging alone carries negative quarters, e.g. 2025.2Q
# -139,938.33백만 -- a "Gain"-named tag holding a loss, read as-is; when a same-quarter
# ...Losses... tag also appears (2025.4Q) it is <0.3% of the Gains tag's magnitude, so the
# dominant tag is taken rather than netted against a tag whose sign convention here is unclear).
# Deliberately EXCLUDES dart_OtherComprehensiveIncomeNetOfTaxGainsLossesOnHedgingInstrument
# (삼성화재 KR0008) -- that is FAIR-VALUE hedge OCI, a different IFRS9 hedge type from cash-flow
# hedge OCI, and every quarter it appears the primary CashFlowHedges tag is ALSO present (0
# cells depend on it; folding it in would only conflate concepts, never fill a real gap).
ACCT_OCI_28_FALLBACK = (
    "dart_GainFromDerivativesHeldForHedging", "dart_GainsValuationDerivativesCashFlowHedge",
    "dart_LossFromDerivativesHeldForHedging", "dart_LossesValuationDerivativesCashFlowHedge",
)
# account_nm fallback (exact match after stripping) for a row tagged "-표준계정코드 미사용-"
# (no account_id at all) -- mirrors _FIN_NM/nm_vals below.  item26: 케이디비생명(KR0072)/
# 푸본현대(KR0083)/코리안리(KR1000) untagged in a few quarters.  item28: 흥국화재(KR0011)/
# KB라이프(KR0099) untagged.
OCI_NM_FALLBACK = {26: ("기타포괄손익-공정가치측정금융자산평가손익",),
                    28: ("위험회피목적파생상품평가손익", "위험회피파생상품평가손익")}

_client = None
_corp_cache: dict[str, str | None] = {}


def _cl():
    global _client
    if _client is None:
        _client = OpenDARTClient.from_settings()
    return _client


def _to_num(x):
    if x in (None, "", "-"):
        return None
    try:
        return float(str(x).replace(",", ""))
    except ValueError:
        return None


def resolve_corp(name):
    """name → DART corp_code (runtime search, cached in-process).  None if not found."""
    if name in _corp_cache:
        return _corp_cache[name]
    queries = [ALIAS.get(name, name), name]
    if name.endswith("생명보험"):
        queries.append(name[:-2])     # 삼성생명보험→삼성생명
    if name.endswith("재보험"):
        queries.append(name[:-3])     # 코리안리재보험→코리안리
    cc = None
    for q in queries:
        try:
            hits = _cl().find_corp_codes_by_name(q)
        except Exception:
            hits = []
        if not hits:
            continue
        exact = [h for h in hits if h["corp_name"] in (q, name)]
        listed = [h for h in hits if h["stock_code"]]
        pick = exact or listed or hits
        cc = pick[0]["corp_code"]
        break
    _corp_cache[name] = cc
    return cc


def _fetch_raw(cc, year, reprt, fs_div, force=False):
    """Cached DART FS-API fetch. The cache under data/dart/_fs_api_cache/ is
    committed (offline PL golden), and it is trusted as canonical once written —
    there is no expiry. That is fine because DART 정정공시 (amended filings) are
    rare, but when one lands the cached JSON for that (corp, year, reprt, fs_div)
    is silently stale forever. To pick up an amendment, delete the matching cache
    file (or run `python scripts/fetch_dart_fs.py --refresh <corp_code> <year>`)
    and rebuild, then commit the refreshed cache alongside the master change.
    `force=True` re-fetches and overwrites regardless of the cached copy.

    Only status=000 (success) responses are persisted. A status:013 ("no data")
    is NOT written to disk -- it's returned for this call but re-checked live next
    time. Confirmed 2026-08-19 (inbox/downloader/20260819T0116Z): a same-quarter
    negative response caches identically to a genuinely-permanent one, but many
    013s are transient (queried the morning after a filing, before DART's FS-API
    had indexed it yet) -- caching those forever turned "not indexed yet" into a
    silent, permanent hole in IFRS17_BS.json. Genuinely permanent gaps (e.g. the
    2023 1Q/2Q coverage void, or non-listed insurers with no XBRL ever) just keep
    re-confirming 013 on every future call -- cheap, and no worse than before."""
    CACHE.mkdir(parents=True, exist_ok=True)
    f = CACHE / f"{cc}_{year}_{reprt}_{fs_div}.json"
    if f.exists() and not force:
        return json.loads(f.read_text(encoding="utf-8"))
    d = _cl()._get("/api/fnlttSinglAcntAll.json",
                   {"corp_code": cc, "bsns_year": str(year), "reprt_code": reprt,
                    "fs_div": fs_div}).json()
    if d.get("status") == "000":
        f.write_text(json.dumps(d, ensure_ascii=False), encoding="utf-8")
    return d


def _parse(d, annual):
    """Parse a fnlttSinglAcntAll response → Tier-1 dict, or None if no income statement."""
    if d.get("status") not in ("000", "013"):
        return None
    vals, is_vals, nm_vals, oci_nm_vals = {}, {}, {}, {}
    # 보험금융 P&L lines some insurers report with '-표준계정코드 미사용-' (no account_id) →
    # collect by NAME for fallback.  OCI uses a distinct name (보험계약자산부채순금융손익), so
    # these exact P&L names don't collide.
    _FIN_NM = ("보험금융수익", "보험금융비용", "재보험금융수익", "재보험금융비용", "보험금융손익")
    for a in d.get("list", []):
        if a.get("sj_div") not in ("IS", "CIS"):
            continue
        aid = a.get("account_id") or ""
        raw = a.get("thstrm_amount") if annual else \
            (a.get("thstrm_add_amount") or a.get("thstrm_amount"))
        v = _to_num(raw)
        if v is None:
            continue
        if aid and aid not in vals:
            vals[aid] = v / 1e6                       # 원 → 백만원
        nm = (a.get("account_nm") or "").replace(" ", "")
        if nm in _FIN_NM and nm not in nm_vals:
            nm_vals[nm] = v / 1e6
        nm_stripped = (a.get("account_nm") or "").strip()
        for oci_item, names in OCI_NM_FALLBACK.items():
            if nm_stripped in names and oci_item not in oci_nm_vals:
                oci_nm_vals[oci_item] = v / 1e6
        for key, pref in IS_PREFIX.items():
            if aid.startswith(pref) and key not in is_vals:
                is_vals[key] = v / 1e6

    def g(aid):
        return vals.get(aid)

    t1 = {}
    for item, aid in ACCT.items():
        v = g(aid)
        if v is not None:
            t1[item] = round(v, 6)
    if 1 not in t1 and 24 not in t1:
        return None                                   # no income statement in this filing
    fi, fe = g(FIN["ins_inc"]), g(FIN["ins_exp"])
    ri, re = g(FIN["re_inc"]), g(FIN["re_exp"])
    # account_nm fallback for insurers whose 보험금융 lines carry no standard account_id
    # (e.g. KB라이프: 보험금융수익 161,082 − 보험금융비용 925,182 = −764,100).
    if fi is None:
        fi = nm_vals.get("보험금융수익")
    if fe is None:
        fe = nm_vals.get("보험금융비용")
    if ri is None:
        ri = nm_vals.get("재보험금융수익")
    if re is None:
        re = nm_vals.get("재보험금융비용")
    if any(x is not None for x in (fi, fe, ri, re)):
        t1[19] = round((fi or 0) - (fe or 0) + (ri or 0) - (re or 0), 6)
    elif nm_vals.get("보험금융손익") is not None:    # only the net line disclosed
        t1[19] = round(nm_vals["보험금융손익"], 6)
    # item17 (투자손익) gross/net consistency: dart_InvestmentIncomeExpenses is GROSS
    # 투자영업손익 for some insurers (영업이익 = 보험손익+투자손익+보험금융손익, e.g. KB라이프) but
    # already NET for others (영업이익 = 보험손익+투자손익, e.g. 한화생명).  When the FS-API 영업이익
    # confirms the gross form, fold 보험금융손익 in so 영업이익 = item1+item17 holds fleet-wide;
    # item18 (투자이익=gross) then derives as item17−item19.  Guarded: only when 1+17+19 closes
    # and 1+17 doesn't, so the already-net insurers are untouched.
    if None not in (t1.get(1), t1.get(17), t1.get(19), t1.get(20)):
        tol = max(0.01 * abs(t1[20]), 200)
        d_net = abs(t1[20] - (t1[1] + t1[17]))
        d_gross = abs(t1[20] - (t1[1] + t1[17] + t1[19]))
        if d_gross <= tol < d_net:
            t1[17] = round(t1[17] + t1[19], 6)        # gross → net
    # item21 (영업외손익): derive as 세전 − 영업이익 (both direct API accounts) so 22=20+21
    # closes exactly; the raw 영업외수익/비용 accounts can miss a sub-line (e.g. 롯데).
    if t1.get(22) is not None and t1.get(20) is not None:
        t1[21] = round(t1[22] - t1[20], 6)
    else:
        oi, oe = g(NONOP_INC), g(NONOP_EXP)
        if oi is not None or oe is not None:
            t1[21] = round((oi or 0) - (oe or 0), 6)
    t1.setdefault(15, 0.0)                            # 기타영업수익: API has no separate acct → 0
    if t1.get(17) is not None and t1.get(19) is not None:
        t1[18] = round(t1[17] - t1[19], 6)
    for k, v in is_vals.items():
        t1[k] = round(v, 6)
    # items 25-31 (기타포괄손익 ~ 총포괄손익) — see ACCT_OCI comment above.
    for item, aid in ACCT_OCI.items():
        v = g(aid)
        if v is not None:
            t1[item] = round(v, 6)
    if 28 not in t1:
        for fid in ACCT_OCI_28_FALLBACK:
            v = g(fid)
            if v is not None:
                t1[28] = round(v, 6)
                break
    for oci_item, val in oci_nm_vals.items():
        t1.setdefault(oci_item, round(val, 6))
    return t1


def tier1_for(name, quarter, code=None):
    """Tier-1 dict (백만원) for one (company, quarter='YYYY.NQ'), or None.
    Basis = CFS for the few 연결-headline groups (BASIS_CFS), else OFS(별도); falls back to
    the other basis if the preferred one has no income statement.  Items 1,15-24 +
    hidden _is_rev/_is_cost/_is_rerev/_is_recost (for Tier-2 item3/8 derivation)."""
    cc = resolve_corp(name)
    if not cc:
        return None
    reprt = REPRT.get(quarter[5:])
    if not reprt:
        return None
    year, annual = quarter[:4], (quarter[5:] == "4Q")
    primary = "CFS" if (code in BASIS_CFS) else "OFS"
    for fs_div in (primary, "CFS" if primary == "OFS" else "OFS"):
        try:
            t1 = _parse(_fetch_raw(cc, year, reprt, fs_div), annual)
        except Exception:
            t1 = None
        if t1:
            return t1
    return None


def _refresh_cache(corp_code: str, year: str) -> int:
    """Re-fetch every cached (reprt × fs_div) for one (corp_code, year), live.
    Use after a DART 정정공시 lands, then rebuild the PL master and commit the
    refreshed cache. Requires OPENDART_API_KEY (live network call).

    Does NOT pre-delete the existing file (2026-08-19 fix): `_fetch_raw` only
    overwrites on a status=000 response, so if this refresh's live call comes
    back 013 (e.g. a transient server hiccup) the previous good cache -- if any
    -- is left in place instead of being deleted with nothing to replace it."""
    n = 0
    for reprt in REPRT.values():
        for fs_div in ("OFS", "CFS"):
            try:
                _fetch_raw(corp_code, year, reprt, fs_div, force=True)
                n += 1
            except Exception as e:  # noqa: BLE001 — report and continue other slices
                print(f"  {reprt}/{fs_div}: {type(e).__name__}: {e}")
    print(f"refreshed {n} cache files for {corp_code} {year}")
    return 0


if __name__ == "__main__":
    if "--refresh" in sys.argv:
        i = sys.argv.index("--refresh")
        try:
            corp_code, year = sys.argv[i + 1], sys.argv[i + 2]
        except IndexError:
            print("usage: fetch_dart_fs.py --refresh <corp_code> <year>", file=sys.stderr)
            raise SystemExit(2)
        raise SystemExit(_refresh_cache(corp_code, year))

    # smoke: validate against the golds
    import openpyxl
    GOLDS = [("삼성화재해상보험", "2025.4Q", "KR0008", "보험손익 breakdown_삼성화재.xlsx"),
             ("메리츠화재해상보험", "2025.4Q", "KR0001", "보험손익 breakdown_메리츠.xlsx"),
             ("삼성생명보험", "2025.4Q", "KR0069", "보험손익 breakdown_삼성생명.xlsx"),
             ("한화생명", "2025.4Q", "KR0068", "보험손익 breakdown_한화생명.xlsx"),
             ("한화생명", "2025.2Q", "KR0068", "보험손익 breakdown_한화생명_2025.2Q.xlsx"),
             ("KB손해보험", "2025.2Q", "KR0010", "보험손익 breakdown_KB.xlsx"),
             ("롯데손해보험", "2024.4Q", "KR0003", "보험손익 breakdown_롯데_2024.xlsx")]
    for nm, q, code, gx in GOLDS:
        wb = openpyxl.load_workbook(gx, data_only=True)
        ws = wb[wb.sheetnames[0]]
        gold = {}
        for row in ws.iter_rows(values_only=True):
            if row and isinstance(row[4], int):
                gold[row[4]] = row[7]
        t1 = tier1_for(nm, q, code) or {}
        line = []
        for it in (1, 17, 19, 20, 22, 23, 24):
            g, v = gold.get(it), t1.get(it)
            ok = g is not None and v is not None and abs(v - g) <= max(1, abs(g) * 0.01)
            line.append(f"{it}:{'OK' if ok else f'{v}vs{g}'}")
        print(f"{nm} {q}: " + "  ".join(line))
