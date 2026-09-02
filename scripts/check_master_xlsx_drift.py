# -*- coding: utf-8 -*-
"""`insurequant_master_tables.xlsx` 의 13개 데이터 시트를 각자의 마스터 JSON 과 셀 단위로 대조한다.

## 왜 있나 (2026-09-02, owner 승인 — 같은 날 사고 2건)

**사고 1.** owner 라이브 QA: NH농협손해보험 2026 기본자본비율 전망이 라이브·마스터는 102.77 인데
xlsx `자본비율전망` 시트만 79.8 이었다. 79.80 = 2026.1Q 기본자본비율, 102.77 = 2026.2Q.
마스터 `kics_forward_capital.json` 은 baseline 을 2026.2Q 로 갱신했는데 **시트만 1Q 기준 옛
산출로 남아 있었다.** 전수 재측정: 38개사 전부, 2090칸 중 **1219칸 stale**.

**사고 2.** owner 가 "그럼 소진율 2종도 stale 하겠네" 라고 지적해 13개 시트를 전수 측정했더니
가설과 결과가 달랐다 — 소진율 2종은 드리프트 0(깨끗)이었고, 아무도 안 보던 **`K-ICS공시`** 가
stale 이었다(변경 33셀 · 추가 121행). 교훈: **어느 시트가 stale 한지 추측하지 말고 전수로 재라.**

**게이트가 왜 못 잡았나.** `PUBLIC_EXPORT_DRIFT`(validate_live_artifacts) 는
마스터 ↔ `public_exports/` 스냅샷만 대조한다. **마스터 ↔ xlsx 를 대조하는 룰이 하나도 없었다** —
xlsx 만 조용히 뒤처져도 RED 가 안 떴다. 이 파일이 그 축이다.

## 설계 결정 3개 (전부 근거가 있다)

### 1. 스키마·평탄화·정규화를 **import 한다. 여기서 다시 적지 않는다.**

시트 목록(`MASTERS`)·행 식별키 집합(`TEXT_COLS`)은 `build_master_xlsx` 에서, 목표행 생성
(`target_rows`)·비교 정규화(`norm`)·키 생성(`key_of`)은 `sync_master_xlsx_sheet` 에서
그대로 가져온다. 평탄화(`FLATTEN`)와 타입 강제(`coerce`)는 여기서 직접 부르지 않는다 —
`target_rows` 가 이미 그 둘을 쓰므로 **동기화가 만드는 목표행과 글자 그대로 같은 것**을 받는다
(비교 기준이 두 벌 생기지 않게 하는 것이 요점이다).
베껴 적으면 빌더가 바뀌는 순간 **검증기가 검증 대상과 다른 스키마를 쓰게 된다**
(상관행렬 재타이핑 금지와 같은 이유). `sync_master_xlsx_sheet` 도 같은 이유로 빌더를
import 한다(그 파일 L13-14).

### 2. 비교 기준 = **동기화 스크립트와 정확히 같은 규칙.** 더 엄해도 안 되고 느슨해도 안 된다.

느슨하면 진짜 값 차이를 놓친다. **엄하면 더 나쁘다** — 게이트가 어떤 도구도 만들 수 없는
상태를 요구하게 되어 영원히 못 고치는 RED 이 된다. 그래서 `norm()` 을 import 해서 쓴다:

  · `'154'`(문자열) vs `154.0`(실수) → 둘 다 `int 154`. 오늘 `K-ICS공시` 에 실제로 있던 형태다.
    이것은 **값 차이가 아니라 셀 타입 차이**다. `coerce()` 가 값이 아닌 열을 전부 문자열로
    만드는데, owner 가 워크북을 Excel 로 열어 저장하면 숫자처럼 보이는 그 텍스트가 숫자로
    바뀐다. 어느 쪽도 값을 안 바꾼다 → 드리프트가 아니다.
    (티커 `'000060'` 은 int 왕복이 깨지므로 문자열로 남는다 — 선행 0 보존)
  · float 는 `%.15g` 로 접는다. xlsx 가 float 를 유효숫자 15자리로 저장하므로
    `168.79000000000002` 는 시트에 `168.79` 로밖에 안 들어간다. 16자리째 차이를 세면
    **영원히 수렴하지 않는 RED** 이 된다(2026-08-21 실측 34칸).
  · 그 위에 **추가 tolerance 는 두지 않는다.** 접은 뒤에는 정확일치를 요구한다 —
    동기화 스크립트가 정확일치를 만들어 내므로 달성 가능하고, 그래야 "값이 다르다"를 안 놓친다.

### 3. 행 식별키 = 값이 아닌 컬럼 전부(`TEXT_COLS` + `항목번호`) — 동기화와 같은 규칙.

`비고` 도 `TEXT_COLS` 에 있으므로 **식별키의 일부**다(실측: `자본비율전망` 의 값 컬럼은
`값` 하나뿐이다). 따라서 비고 문구가 바뀐 행은 EDIT 이 아니라 **ROW_MISSING + ROW_EXTRA 쌍**
으로 나온다 — 오늘 `자본비율전망` 에서 169행이 그렇게 잡혔고 동기화도 그렇게 반영했다.
키를 다르게 잡으면 게이트와 동기화가 갈라져서, 동기화가 방금 만든 상태를 게이트가 계속
드리프트라고 부르게 된다. **키 규칙은 반드시 같아야 한다.**

## `요약` 시트 — 검사하되 **행수만** 본다 (근거)

`요약` 은 파생 시트다(시트명 · 마스터 파일 · 행수 · 설명). 이 중

  · **행수**는 기계가 유지한다(`sync_master_xlsx_sheet.py` L292-303 이 실측으로 갱신) →
    틀리면 owner 가 보는 xlsx 에 틀린 수가 인쇄된다 → **검사한다.**
  · **설명**은 다른 레인이 손으로 고쳐 둔 문구다(같은 파일 L21-22, L271-272 가 "손대지 않는다"
    고 명시) → 기계가 정본을 갖고 있지 않다 → **검사하지 않는다.** 검사하면 owner/레인의
    정당한 손질이 매번 RED 이 된다.

## MASTERS 밖 시트

수기 시트(피벗 등)를 넣는 것은 허용된 설계다(`sync_master_xlsx_sheet.py` L8-10). 그래서 RED 이
아니라 **census YELLOW** 로 센다 — "검사 안 되는 축"이 조용히 존재하는 것만은 막는다.

## 사용

    C:/Users/sangwook.cho/venvs/insurequant/Scripts/python.exe scripts/check_master_xlsx_drift.py
    ... --sheet 자본비율전망      # 한 시트만
    exit 2 = RED 있음

게이트 배선: `scripts/validate_data_contract.py` CHECK 8 `check_master_xlsx`
(→ `run_gate` → `scripts/prepush_check.py` §1 → `.githooks/pre-push`).
"""
from __future__ import annotations

import re
import sys
import zipfile
from collections import Counter
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "scripts"))

from build_master_xlsx import MASTERS, TEXT_COLS                      # noqa: E402
from sync_master_xlsx_sheet import key_of, norm, target_rows          # noqa: E402

XLSX = REPO / "insurequant_master_tables.xlsx"
SUMMARY_SHEET = "요약"
SUMMARY_TOTAL_LABEL = "합계"
MAX_SAMPLES = 5              # 1219칸 같은 대량 케이스에서 출력이 폭발하지 않게
_FORMULA_TAG = re.compile(rb"<f[\s/>]")   # `<f>` · `<f t="shared" .../>` · `<f t="array" ...>`

RULES = (
    "MASTER_XLSX_FILE_MISSING",
    "MASTER_XLSX_UNREADABLE",
    "MASTER_XLSX_FORMULA_PRESENT",
    "MASTER_XLSX_SHEET_MISSING",
    "MASTER_XLSX_COLUMN_MISMATCH",
    "MASTER_XLSX_MASTER_UNREADABLE",
    "MASTER_XLSX_KEY_AMBIGUOUS",
    "MASTER_XLSX_ROW_MISSING",
    "MASTER_XLSX_ROW_EXTRA",
    "MASTER_XLSX_DRIFT",
    "MASTER_XLSX_SUMMARY_ROWCOUNT",
    "MASTER_XLSX_SUMMARY_SHEET_MISSING",
    "MASTER_XLSX_UNTRACKED_SHEET",
)


def _f(severity, sheet, rule, message, count=1):
    """`count` = 이 finding 이 대표하는 셀·행 수(집계는 stat 이 쓴다. 메시지 문자열을 다시
    파싱해서 세지 않는다 — 그러면 문구를 고치는 순간 집계가 조용히 틀어진다)."""
    return {"severity": severity, "sheet": sheet, "rule": rule,
            "message": message, "count": count}


# --------------------------------------------------------------------------- I/O
def read_workbook(path: Path = XLSX) -> tuple[dict, list]:
    """`{sheet: (header, rows)}` + 치명적 finding.

    **읽기 전용으로만 연다** (`data_only=True, read_only=True`). openpyxl 로 load+save 하면
    다른 시트의 수식 캐시가 통째로 날아간다(memory `project_master_xlsx_formula_cache`) —
    이 게이트는 워크북을 절대 쓰지 않는다.
    """
    if not path.exists():
        return {}, [_f("RED", "-", "MASTER_XLSX_FILE_MISSING",
                       f"{path.name} 이 없다 — owner 가 받아 보는 마스터 워크북 자체가 사라졌다")]
    # 수식 존재 여부는 워크북을 열기 전에 raw xml 로 센다(0.1초). 수식이 있으면
    # `data_only=True` 가 **캐시값**을 읽으므로 비교가 조용히 무의미해진다 —
    # `sync_master_xlsx_sheet.py` 도 같은 이유로 수식이 하나라도 있으면 실행을 거부한다.
    # `<f>` 만 세면 안 된다: 채우기로 만든 **공유수식**은 `<f t="shared" si="0"/>` 로 쓰이고
    # 배열수식은 `<f t="array" ref=...>` 다. 셋 다 수식이므로 여는 태그 전체를 센다.
    try:
        with zipfile.ZipFile(path) as z:
            n_formula = sum(len(_FORMULA_TAG.findall(z.read(n))) for n in z.namelist()
                            if n.startswith("xl/worksheets/sheet"))
    except Exception as e:
        return {}, [_f("RED", "-", "MASTER_XLSX_UNREADABLE",
                       f"{path.name} 을 zip 으로 못 연다 ({type(e).__name__}: {e}) — "
                       f"깨진 파일은 없는 파일과 다르다")]
    out_findings = []
    if n_formula:
        out_findings.append(_f(
            "RED", "-", "MASTER_XLSX_FORMULA_PRESENT",
            f"워크북에 수식 {n_formula}개가 있다 — data_only 읽기는 **캐시값**을 보게 되므로 "
            f"이 축의 비교가 조용히 무의미해진다. sync_master_xlsx_sheet.py 도 같은 이유로 "
            f"실행을 거부한다(그 파일 L138-143). 수식을 없애거나 워크북을 재생성해라"))

    from openpyxl import load_workbook
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        out_findings.append(_f("RED", "-", "MASTER_XLSX_UNREADABLE",
                               f"{path.name} 을 openpyxl 로 못 읽는다 "
                               f"({type(e).__name__}: {e})"))
        return {}, out_findings
    sheets = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if header is None:
                sheets[name] = ([], [])
                continue
            header = list(header)
            ncol = len(header)
            rows = []
            for r in it:
                r = list(r)[:ncol] + [None] * max(0, ncol - len(r))
                rows.append(r)
            # Excel 이 남긴 꼬리 빈 행은 데이터가 아니다(전 칸 None 인 뒷줄만 잘라낸다).
            while rows and all(v is None or v == "" for v in rows[-1]):
                rows.pop()
            sheets[name] = (header, rows)
    finally:
        wb.close()
    return sheets, out_findings


# --------------------------------------------------------------------------- pure compare
def compare_sheet(sheet: str, cols: list, tgt: list, cur_header: list, cur: list) -> list:
    """마스터 목표행(`tgt`) 과 시트 실제행(`cur`) 을 셀 단위로 대조한다 — 순수 함수.

    디스크를 안 건드리므로 변이시험이 워크북을 재저장할 필요가 없다(그게 이 분리의 목적이다).
    """
    if list(cur_header) != list(cols):
        return [_f("RED", sheet, "MASTER_XLSX_COLUMN_MISMATCH",
                   f"헤더가 빌더 스키마와 다르다\n         시트  : {list(cur_header)}"
                   f"\n         마스터: {list(cols)}")]

    ncol = len(cols)
    key_idx = [i for i, c in enumerate(cols) if c in TEXT_COLS or c == "항목번호"]
    val_idx = [i for i in range(ncol) if i not in key_idx]
    if not key_idx or not val_idx:
        return [_f("RED", sheet, "MASTER_XLSX_KEY_AMBIGUOUS",
                   f"식별열/값열을 나누지 못했다(cols={cols}) — build_master_xlsx.TEXT_COLS 를 "
                   f"갱신해라. 이대로 두면 이 시트의 값 비교가 통째로 건너뛰어진다")]

    tgt_keys = [key_of(r, cols, key_idx) for r in tgt]
    cur_keys = [key_of(r, cols, key_idx) for r in cur]
    tmap, tdup = {}, 0
    for k, r in zip(tgt_keys, tgt):
        if k in tmap:
            tdup += 1
        tmap[k] = r
    if tdup:
        return [_f("RED", sheet, "MASTER_XLSX_KEY_AMBIGUOUS",
                   f"마스터 쪽에서 행 식별키 {[cols[i] for i in key_idx]} 가 유일하지 않다"
                   f"(중복 {tdup}행) — 셀 비교가 성립하지 않는다. 식별열을 늘려라 "
                   f"(자본성증권발행현황이 2026-09-01 에 그랬다: 증권명이 키에서 빠져 있었다)")]
    cmap = {}
    for k, r in zip(cur_keys, cur):
        cmap.setdefault(k, r)

    out = []
    missing = [k for k in tgt_keys if k not in cmap]
    # 시트에만 있는 행 = ① 마스터에 없는 키 + ② **키가 중복된 행의 초과분**.
    # ②를 안 세면 조용히 새는 구멍이 생긴다: 행이 통째로 복제되면 그 키는 마스터에 있으므로
    # missing 에도 extra 에도 안 걸리고, 값 비교는 첫 행만 보므로 전부 일치로 끝난다.
    # 그러면 owner 워크북에 중복 행이 있는 채로 게이트가 초록을 찍는다.
    dup_surplus = [k for k, n in Counter(cur_keys).items() for _ in range(n - 1)]
    extra = [k for k in cur_keys if k not in tmap] + dup_surplus
    if missing:
        out.append(_f("RED", sheet, "MASTER_XLSX_ROW_MISSING",
                      f"마스터에 있는데 시트에 없는 행 {len(missing)}건 "
                      f"(예: {missing[:MAX_SAMPLES]}) — 기대 그리드는 마스터다. "
                      f"python scripts/sync_master_xlsx_sheet.py \"{sheet}\"", len(missing)))
    if extra:
        out.append(_f("RED", sheet, "MASTER_XLSX_ROW_EXTRA",
                      f"시트에만 있는 행 {len(extra)}건"
                      + (f" (그중 중복 복제 {len(dup_surplus)}건)" if dup_surplus else "")
                      + f" (예: {extra[:MAX_SAMPLES]}) — "
                      f"마스터에서 사라졌거나 복제된 행이 워크북에 남아 있다. "
                      f"python scripts/sync_master_xlsx_sheet.py \"{sheet}\"", len(extra)))

    drift, samples = 0, []
    for k in tgt_keys:
        crow = cmap.get(k)
        if crow is None:
            continue
        trow = tmap[k]
        for i in val_idx:
            a, b = norm(crow[i]), norm(trow[i])
            if a != b:
                drift += 1
                if len(samples) < MAX_SAMPLES:
                    samples.append(f"{k}|{cols[i]}: 시트={crow[i]!r} vs 마스터={trow[i]!r}")
    if drift:
        out.append(_f("RED", sheet, "MASTER_XLSX_DRIFT",
                      f"값이 다른 셀 {drift}건 / {len(tgt)}행 — " + " / ".join(samples) +
                      (f" (+{drift - len(samples)}건 더)" if drift > len(samples) else "") +
                      f". 마스터가 갱신됐는데 시트 동기화가 밀렸다: "
                      f"python scripts/sync_master_xlsx_sheet.py \"{sheet}\"", drift))
    return out


def compare_summary(sheets: dict, declared: list) -> list:
    """`요약` 시트의 **행수만** 검사한다 (설명 열은 손으로 관리한다 — 모듈 docstring 참조)."""
    if SUMMARY_SHEET not in sheets:
        return [_f("RED", SUMMARY_SHEET, "MASTER_XLSX_SUMMARY_SHEET_MISSING",
                   f"'{SUMMARY_SHEET}' 색인 시트가 없다")]
    _title, rows = sheets[SUMMARY_SHEET]
    # 요약 시트는 1행=제목 · 2행=빈줄 · 3행=헤더 · 4행부터 본문 + 마지막 '합계' 다.
    # read_workbook 이 1행을 떼어 갔으므로 rows 에는 빈줄·헤더가 남는데, 아래는 **선언된
    # 시트명으로만 조회**하므로 그 두 줄은 조회되지 않는다(따로 걸러낼 필요가 없다).
    body = {}
    for r in rows:
        name = r[0] if r else None
        if not name or name == SUMMARY_TOTAL_LABEL:
            continue
        body[name] = r[2] if len(r) > 2 else None

    out, total = [], 0
    for _fn, sheet, _desc in declared:
        actual = len(sheets[sheet][1]) if sheet in sheets else None
        if actual is None:
            continue                       # 시트 자체 부재는 MASTER_XLSX_SHEET_MISSING 이 낸다
        total += actual
        if sheet not in body:
            out.append(_f("RED", SUMMARY_SHEET, "MASTER_XLSX_SUMMARY_SHEET_MISSING",
                          f"'{sheet}' 시트가 워크북에 있는데 요약 색인에 줄이 없다 — "
                          f"owner 는 색인만 보고 시트 유무를 판단한다"))
            continue
        listed = body[sheet]
        if norm(listed) != actual:
            out.append(_f("RED", SUMMARY_SHEET, "MASTER_XLSX_SUMMARY_ROWCOUNT",
                          f"'{sheet}' 행수: 요약={listed!r} vs 실제={actual} — "
                          f"owner 가 보는 색인이 틀린 수를 인쇄한다"))
    listed_total = next((r[2] for r in rows if r and r[0] == SUMMARY_TOTAL_LABEL), None)
    if listed_total is not None and norm(listed_total) != total:
        out.append(_f("RED", SUMMARY_SHEET, "MASTER_XLSX_SUMMARY_ROWCOUNT",
                      f"'{SUMMARY_TOTAL_LABEL}' 행수: 요약={listed_total!r} vs 실제 합계={total}"))
    return out


# --------------------------------------------------------------------------- driver
def scan(path: Path = XLSX, only: str | None = None,
         sheets: dict | None = None) -> tuple[list, dict]:
    """전 시트 대조. `(findings, stat)`. 워크북을 **읽기만** 한다.

    `sheets` 를 주면 디스크를 아예 안 읽는다 — 변이시험이 워크북을 **재저장하지 않고**
    이 드라이버를 그대로 태우기 위한 자리다(memory `project_master_xlsx_formula_cache`:
    openpyxl load+save 는 다른 시트의 수식 캐시를 통째로 날린다. 그래서 변이는 전부
    메모리 안에서 한다).
    """
    declared = [m for m in MASTERS if only is None or m[1] == only]
    if sheets is None:
        sheets, findings = read_workbook(path)
    else:
        findings = []
    stat = {"sheets_declared": len(declared), "sheets_compared": 0,
            "rows_compared": 0, "cells_drifted": 0}
    if not sheets:
        return findings, stat

    for json_file, sheet, _desc in declared:
        if sheet not in sheets:
            findings.append(_f("RED", sheet, "MASTER_XLSX_SHEET_MISSING",
                               f"{json_file} 의 시트가 워크북에 없다 — 이 마스터가 owner 워크북에서 "
                               f"통째로 빠져 있다"))
            continue
        try:
            cols, tgt = target_rows(json_file)
        except Exception as e:
            findings.append(_f("RED", sheet, "MASTER_XLSX_MASTER_UNREADABLE",
                               f"마스터 {json_file} 에서 목표행을 못 만든다 "
                               f"({type(e).__name__}: {e}) — 대조 기준이 없으면 이 시트는 "
                               f"조용히 무검사가 된다"))
            continue
        header, cur = sheets[sheet]
        got = compare_sheet(sheet, cols, tgt, header, cur)
        findings.extend(got)
        stat["sheets_compared"] += 1
        stat["rows_compared"] += len(tgt)
        stat["cells_drifted"] += sum(g["count"] for g in got
                                     if g["rule"] == "MASTER_XLSX_DRIFT")

    if only is None:
        findings.extend(compare_summary(sheets, declared))
        known = {m[1] for m in MASTERS} | {SUMMARY_SHEET}
        for name in sheets:
            if name not in known:
                findings.append(_f(
                    "YELLOW", name, "MASTER_XLSX_UNTRACKED_SHEET",
                    f"워크북에 '{name}' 시트가 있는데 build_master_xlsx.MASTERS 에 없다 — "
                    f"수기 시트는 허용된 설계지만(sync_master_xlsx_sheet.py L8-10) "
                    f"**이 시트는 어떤 검사도 안 받는다**. 마스터가 생겼으면 MASTERS 에 등재해라"))
    return findings, stat


def main(argv=None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    only = None
    if "--sheet" in argv:
        only = argv[argv.index("--sheet") + 1]
    findings, stat = scan(only=only)
    print("=" * 78)
    print("MASTER XLSX DRIFT  (insurequant_master_tables.xlsx  vs  루트 마스터 JSON)")
    print("=" * 78)
    print(f"  선언 시트 {stat['sheets_declared']} · 대조 시트 {stat['sheets_compared']} · "
          f"대조 행 {stat['rows_compared']} · 드리프트 셀 {stat['cells_drifted']}")
    red = [f for f in findings if f["severity"] == "RED"]
    for f in findings:
        print(f"  {f['severity']:6s} [{f['sheet']}] {f['rule']}")
        print(f"         {f['message']}")
    if not findings:
        print("  (clean)")
    print("=" * 78)
    print(f"SUMMARY  RED={len(red)}  YELLOW={len(findings) - len(red)}")
    return 2 if red else 0


if __name__ == "__main__":
    raise SystemExit(main())
