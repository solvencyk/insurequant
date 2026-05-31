#!/usr/bin/env python3
"""미래에셋생명 (KR0079): parse Fact Sheet CSM sheet -> 신계약 CSM series.

미래에셋생명 does NOT disclose a 신계약 CSM 배수 (multiple) in its fact sheet.
We therefore capture the disclosed 신계약 CSM amount (total + by-product), in 억원,
per quarter FY2023.1Q ~ FY2026.1Q.

Source layout (CSM sheet): row5 = year (FY/2026/2025/2024/2023/2022),
row6 = period (Q1/FY/Q4/Q3/Q2/Q1 ...). Values in BN KRW (십억원) -> x10 = 억원.
신계약CSM = CSM!row8 (total); by-product rows 9 (보장성), 10 (일반저축), 11 (변액저축).

We read EVERY downloaded fact sheet and merge; the newest carries the most quarters
as comparative columns, so its (possibly restated) values win. We record which file
each quarter's value came from.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
KR_DIR = "KR0079_미래에셋생명"
OUT = ROOT / "data" / "ir" / "series" / f"{KR_DIR}.json"

# target window
WINDOW = {f"{y}.{q}" for y in range(2023, 2027) for q in ("1Q", "2Q", "3Q", "4Q")}
WINDOW = {p for p in WINDOW if not (p.startswith("2026") and p != "2026.1Q")}


def find_csm_sheet(wb):
    for sh in wb.worksheets:
        if sh.title.strip().upper() == "CSM":
            return sh
    # fallback: a sheet whose B1 == 'CSM'
    for sh in wb.worksheets:
        v = sh.cell(1, 2).value
        if isinstance(v, str) and v.strip() == "CSM":
            return sh
    return None


def build_col_period(ws):
    """Find the year row + quarter row in the CSM movement block, map col->period."""
    year_row = q_row = None
    for r in range(1, 12):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # year row: has multiple 4-digit years among 2018..2027
        yrs = [v for v in vals if isinstance(v, (int, float, str))
               and re.fullmatch(r"20\d{2}", str(v).strip().split(".")[0] if v is not None else "")]
        if year_row is None and len(yrs) >= 2:
            year_row = r
        # quarter row: has >=3 of Q1..Q4 / FY
        qs = sum(1 for v in vals if isinstance(v, str) and v.strip() in ("Q1", "Q2", "Q3", "Q4", "FY", "1", "2", "3", "4"))
        if q_row is None and year_row is not None and r > year_row and qs >= 3:
            q_row = r
    if year_row is None or q_row is None:
        return {}
    # forward-fill year across columns
    col_year = {}
    cur = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(year_row, c).value
        s = str(v).strip().split(".")[0] if v is not None else ""
        if re.fullmatch(r"20\d{2}", s):
            cur = s
        col_year[c] = cur
    col_period = {}
    for c in range(1, ws.max_column + 1):
        q = ws.cell(q_row, c).value
        yr = col_year.get(c)
        if not yr:
            continue
        qs = str(q).strip() if q is not None else ""
        # accept Q1..Q4 (skip 'FY' cumulative, and skip the YoY duplicate Q1 block by
        # preferring the per-FY breakdown columns); map numeric 1..4 too
        m = {"Q1": "1Q", "Q2": "2Q", "Q3": "3Q", "Q4": "4Q",
             "1": "1Q", "2": "2Q", "3": "3Q", "4": "4Q"}.get(qs)
        if m:
            per = f"{yr}.{m}"
            # keep first occurrence (the dedicated per-year breakdown columns appear
            # after the YoY pair; first hit for current-year Q1 is the YoY one, which
            # is identical value, so fine)
            col_period.setdefault(c, per)
    return col_period


def row_with_label(ws, label):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() == label:
                return c.row
    return None


BN_TO_EOK = 10.0  # 십억원 -> 억원


def parse_factsheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_csm_sheet(wb)
    if ws is None:
        return {}
    col_period = build_col_period(ws)
    if not col_period:
        return {}
    r_total = row_with_label(ws, "신계약CSM")
    r_prot = row_with_label(ws, "보장성상품(일반,변액)")
    r_gen = row_with_label(ws, "일반저축성")
    r_var = row_with_label(ws, "변액저축성")
    out = {}
    for c, per in col_period.items():
        if per not in WINDOW:
            continue

        def val(r):
            if not r:
                return None
            v = ws.cell(r, c).value
            return round(float(v) * BN_TO_EOK, 1) if isinstance(v, (int, float)) else None

        tot = val(r_total)
        if tot is None:
            continue
        rec = {"nb_csm_eok": tot}
        bp = {}
        for nm, r in (("보장성", r_prot), ("일반저축", r_gen), ("변액저축", r_var)):
            x = val(r)
            if x is not None:
                bp[nm] = x
        if bp:
            rec["nb_csm_by_product_eok"] = bp
        out[per] = rec
    return out


def main():
    files = sorted(ROOT.glob("data/ir/FY*/" + KR_DIR + "/*FactSheet*.xlsx"))
    print(f"{len(files)} fact sheets")
    merged = {}
    src = {}
    # process oldest->newest so newest (restated) wins
    for f in files:
        try:
            ser = parse_factsheet(f)
        except Exception as e:  # noqa: BLE001
            print(f"  parse fail {f.name}: {e}")
            continue
        for per, rec in ser.items():
            merged[per] = rec
            src[per] = f.name
        print(f"  {f.relative_to(ROOT)} -> {sorted(ser)}")
    for per in merged:
        merged[per]["source_file"] = src[per]
    payload = {
        "company": "미래에셋생명",
        "kr_code": "KR0079",
        "sector": "Life",
        "publisher": "미래에셋생명 IR 자료실 (life.miraeasset.com) — quarterly Fact Sheet (Excel)",
        "metric": "신계약 CSM (억원) — 미래에셋생명 Fact Sheet 'CSM' sheet row '신계약CSM' (별도 추정; 십억원 단위 x10)",
        "multiple_disclosed": False,
        "multiple_note": "미래에셋생명 Fact Sheet does NOT disclose a 신계약 CSM 배수 (월납월초/APE multiple). Only 신계약 CSM amount (total + by-product: 보장성/일반저축/변액저축) and 신계약 APE / 초회보험료 are disclosed. Captured 신계약 CSM amount per brief.",
        "ir_page": "https://life.miraeasset.com/micro/company/PC-HO-060401-000000.do",
        "source_note": "Latest fact sheet (2026 Q1) carries 2022~2026 quarters as comparative columns; per-quarter source_file recorded. Life insurer => FY == calendar year.",
        "units": {"nb_csm_eok": "KRW 100M (억원)"},
        "series": dict(sorted(merged.items())),
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nwrote {OUT}")
    for per, rec in sorted(merged.items()):
        bp = rec.get("nb_csm_by_product_eok", {})
        print(f"  {per}: 신계약CSM={rec['nb_csm_eok']}억  by={bp}  <- {rec['source_file']}")


if __name__ == "__main__":
    main()
