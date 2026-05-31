#!/usr/bin/env python3
"""Dump 미래에셋생명 CSM + INITIAL/NB sheets fully to understand layout & periods."""
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
f = ROOT / "data/ir/FY2026_Q1/KR0079_미래에셋생명/2026 Q1 FactSheet_Kr.xlsx"
wb = openpyxl.load_workbook(f, data_only=True)

# global search for 배수 / multiple / x
print("== global search for '배수'/'multiple' ==")
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value, str) and ("배수" in c.value or "multiple" in c.value.lower()):
                print(f"  {sh.title}!({c.row},{c.column}) {c.value!r}")


def dump(sheet, rmax=20, cmax=30):
    ws = wb[sheet]
    print(f"\n===== {sheet} (showing r1..{rmax}, c1..{cmax}) =====")
    for r in range(1, min(rmax, ws.max_row) + 1):
        cells = []
        for c in range(1, min(cmax, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                sv = str(v)
                if len(sv) > 14:
                    sv = sv[:14]
                cells.append(f"c{c}={sv}")
        if cells:
            print(f" r{r}: " + " | ".join(cells))


dump("CSM", 20, 40)
dump("INITIAL", 22, 30)
dump("NB", 25, 30)
dump("APE", 16, 30)
