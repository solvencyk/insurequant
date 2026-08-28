import sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

wb = openpyxl.load_workbook("insurequant_master_tables.xlsx", read_only=True, data_only=False)
ws = wb["손익분해PL"]
print("max_row:", ws.max_row)
# find header
header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("header:", header)
idx_item = header.index("항목번호")
idx_name = header.index("항목명")
count32 = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[idx_item] == 32:
        count32 += 1
print("rows with 항목번호==32:", count32)
wb.close()
