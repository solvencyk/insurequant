"""Show a few K-ICS공시 rows with python types + number formats."""
import io, sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
REPO = Path(__file__).resolve().parents[2]
wb = load_workbook(REPO / "insurequant_master_tables.xlsx", data_only=False)
ws = wb["K-ICS공시"]
hdr = [c.value for c in ws[1]]
print(hdr)
want = {("KR0050", "2023.2Q"), ("KR0097", "2023.2Q"), ("KR0071", "2023.1Q")}
shown = 0
for r in range(2, ws.max_row + 1):
    code = ws.cell(row=r, column=1).value
    q = ws.cell(row=r, column=7).value
    if (code, q) in want and shown < 12:
        vals = [(ws.cell(row=r, column=c).value, type(ws.cell(row=r, column=c).value).__name__,
                 ws.cell(row=r, column=c).number_format) for c in (5, 8, 9)]
        print(f"  row{r} {code} {q} item/값/값_적용후 = {vals}")
        shown += 1
print("--- 요약 sheet ---")
idx = wb["요약"]
for r in range(1, idx.max_row + 1):
    print("  ", [idx.cell(row=r, column=c).value for c in range(1, 5)])
