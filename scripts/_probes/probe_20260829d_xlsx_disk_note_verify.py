# -*- coding: utf-8 -*-
"""Independent re-verification (fresh session, before commit): read the 기본자본소진율
sheet DIRECTLY off insurequant_master_tables.xlsx on disk (not the FLATTEN() function in
memory) and confirm the 13-row over-100 비고 patch actually landed in the saved file."""
from __future__ import annotations

import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
REPO = Path(__file__).resolve().parents[2]
from openpyxl import load_workbook  # noqa: E402

wb = load_workbook(REPO / "insurequant_master_tables.xlsx", data_only=False)
ws = wb["기본자본소진율"]
hdr = [c.value for c in ws[1]]
col = {name: i + 1 for i, name in enumerate(hdr)}

pct_rows = []
for r in range(2, ws.max_row + 1):
    item = ws.cell(row=r, column=col["항목명"]).value
    if item in ("기본자본 소진율", "기본자본 소진율(엄격)"):
        pct_rows.append({
            "row": r,
            "company": ws.cell(row=r, column=col["원수사명"]).value,
            "item": item,
            "val": ws.cell(row=r, column=col["값"]).value,
            "note": ws.cell(row=r, column=col["비고"]).value or "",
        })

print(f"소진율 rows on disk: {len(pct_rows)} (expect 78)")
empty = [r for r in pct_rows if not r["note"]]
print(f"empty 비고: {len(empty)} (expect 0)")
over100 = [r for r in pct_rows if (r["val"] or 0) > 100.0]
print(f"값>100 rows: {len(over100)} (expect 13)")
missing_clause = [r for r in over100 if "100%초과는 파싱오류 아님" not in r["note"]]
print(f"over-100 rows MISSING the clause: {len(missing_clause)} (expect 0)")
under100_with_clause = [r for r in pct_rows if (r["val"] or 0) <= 100.0
                         and "100%초과는 파싱오류 아님" in r["note"]]
print(f"<=100 rows wrongly carrying the clause: {len(under100_with_clause)} (expect 0)")
missing_basis = [r for r in pct_rows if "SCR×15%" not in r["note"] and "SCR×10%" not in r["note"]]
print(f"rows missing the SCR basis explanation: {len(missing_basis)} (expect 0)")

print("\nthe 13 over-100 rows:")
for r in sorted(over100, key=lambda x: (x["company"], x["item"])):
    print(f"  {r['company']:12s} {r['item']:22s} {r['val']}")

print("\nDONE")
