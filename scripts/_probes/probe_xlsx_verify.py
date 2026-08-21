"""Independent post-sync check: new xlsx vs backup + vs master JSON."""
import io, json, sys, zipfile
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO / "scripts"))
from build_master_xlsx import coerce, NUMERIC_COLS

BAK = Path(sys.argv[1])
NEW = REPO / "insurequant_master_tables.xlsx"

def rows(ws):
    return [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
def norm(v):
    if v is None or v == "": return None
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v

a = load_workbook(BAK); b = load_workbook(NEW)
print("sheets same:", a.sheetnames == b.sheetnames, b.sheetnames)
for t in b.sheetnames:
    ra, rb = [[norm(v) for v in r] for r in rows(a[t])], [[norm(v) for v in r] for r in rows(b[t])]
    tag = "SYNCED" if t == "K-ICS공시" else ("요약" if t == "요약" else "untouched")
    same = ra == rb
    print(f"  {t:14s} {len(ra):>6} -> {len(rb):>6} rows  identical={same}  [{tag}]")
    if not same and tag == "untouched":
        print("     !!! 손대지 않아야 할 시트가 변함")

with zipfile.ZipFile(NEW) as z:
    print("parts:", len(z.namelist()))
nf = sum(1 for ws in b.worksheets for r in ws.iter_rows() for c in r
         if isinstance(c.value, str) and c.value.startswith("="))
print("formulas in new file:", nf)

data = json.loads((REPO / "kics_disclosure.json").read_text(encoding="utf-8"))
df = coerce(pd.DataFrame(data))
ws = b["K-ICS공시"]
cols = [c.value for c in ws[1]]
mis = 0
for i, rec in enumerate(df.to_dict("records")):
    for j, c in enumerate(cols):
        v = rec[c]
        exp = None if (v is None or v is pd.NA or (isinstance(v, float) and pd.isna(v))) else (
            (int(float(v)) if float(v).is_integer() else float(v)) if (c in NUMERIC_COLS or c == "항목번호") else str(v))
        if norm(ws.cell(row=i + 2, column=j + 1).value) != norm(exp):
            mis += 1
print(f"K-ICS공시 vs kics_disclosure.json 셀 불일치: {mis} (0이어야 정상, {len(df)}행 × {len(cols)}열)")
