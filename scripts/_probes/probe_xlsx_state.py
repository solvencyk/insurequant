"""Inspect insurequant_master_tables.xlsx: sheets, dims, header, formula census."""
import io, sys
from pathlib import Path
from openpyxl import load_workbook
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
REPO = Path(__file__).resolve().parents[2]
wb = load_workbook(REPO / "insurequant_master_tables.xlsx", data_only=False)
print("sheets:", wb.sheetnames)
for ws in wb.worksheets:
    formulas = 0
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                formulas += 1
    hdr = [c.value for c in ws[1]]
    print(f"  {ws.title!r}: rows={ws.max_row} cols={ws.max_column} formulas={formulas}")
    print(f"      header={hdr}")
