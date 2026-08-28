# -*- coding: utf-8 -*-
"""Read-only post-sync sanity check on insurequant_master_tables.xlsx (no save)."""
from __future__ import annotations

import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO = Path(__file__).resolve().parents[2]
from openpyxl import load_workbook  # noqa: E402

wb = load_workbook(REPO / "insurequant_master_tables.xlsx", data_only=False)
print("sheet order:", wb.sheetnames)

n_formula = sum(
    1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
    if isinstance(c.value, str) and c.value.startswith("=")
)
print("total formulas across workbook:", n_formula, "(must be 0)")

idx = wb["요약"]
print("\n요약 sheet contents:")
for r in range(1, idx.max_row + 1):
    vals = [idx.cell(row=r, column=c).value for c in range(1, 5)]
    print(f"  row{r}: {vals}")

for name in ("기본자본소진율", "보완자본소진율", "자본비율전망"):
    ws = wb[name]
    print(f"\n=== {name}: {ws.max_row - 1} data rows, {ws.max_column} cols, "
          f"freeze_panes={ws.freeze_panes}, auto_filter={ws.auto_filter.ref} ===")
    hdr = [ws.cell(row=1, column=c) for c in range(1, ws.max_column + 1)]
    print("  header values:", [c.value for c in hdr])
    print("  header font:", hdr[0].font.bold, hdr[0].font.color.rgb if hdr[0].font.color else None,
          "fill:", hdr[0].fill.fgColor.rgb if hdr[0].fill and hdr[0].fill.fgColor else None)
    r2 = [ws.cell(row=2, column=c) for c in range(1, ws.max_column + 1)]
    print("  row2 values:", [c.value for c in r2])
    print("  row2 font bold(should be False):", r2[0].font.bold, "number_format(값 col):",
          ws.cell(row=2, column=7).number_format)
    widths = {ws.cell(row=1, column=c).value: ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width
              for c in range(1, ws.max_column + 1)}
    print("  col widths:", widths)

print("\nDONE")
