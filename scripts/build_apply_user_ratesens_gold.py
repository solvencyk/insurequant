# -*- coding: utf-8 -*-
"""Capture + apply owner hand-filled 금리민감도 (rate-sensitivity) rows.

The owner added rows for life insurers the extractor missed (extract_kics_rate_
sensitivity.py gap). build mode captures owner-only rows from the review xlsx
into data/_gold/user_rate_sensitivity_rows.json; apply mode upserts them into
kics_rate_sensitivity.json so they survive a re-extract.

Key = (원보험사코드, 공시분기, 경과조치여부, measure구분). Owner rows win.

Usage:
  PYTHONIOENCODING=utf-8 python scripts/build_apply_user_ratesens_gold.py build
  PYTHONIOENCODING=utf-8 python scripts/build_apply_user_ratesens_gold.py apply
"""
from __future__ import annotations
import json
import os
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
REPO = Path(__file__).resolve().parents[1]
RS_JSON = REPO / "kics_rate_sensitivity.json"
GOLD = REPO / "data" / "_gold" / "user_rate_sensitivity_rows.json"
COLS = ["-100bp", "-50bp", "base", "+50bp", "+100bp"]
KEYF = ("원보험사코드", "공시분기", "경과조치여부", "measure구분")


def _key(r: dict) -> tuple:
    return tuple(r.get(k) for k in KEYF)


def _num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def build(xlsx_path: str) -> int:
    rs = json.loads(RS_JSON.read_text(encoding="utf-8"))
    jkeys = {_key(r) for r in rs}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["금리민감도"]
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {h: i for i, h in enumerate(hdr)}
    new_rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        rec = {h: r[idx[h]] for h in hdr}
        if _key(rec) in jkeys:
            continue
        for c in COLS:
            rec[c] = _num(rec.get(c))
        new_rows.append(rec)
    wb.close()
    GOLD.parent.mkdir(parents=True, exist_ok=True)
    GOLD.write_text(json.dumps(new_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"rate-sens gold built: {len(new_rows)} owner rows -> {GOLD}")
    return 0


def apply() -> int:
    rs = json.loads(RS_JSON.read_text(encoding="utf-8"))
    gold = json.loads(GOLD.read_text(encoding="utf-8"))
    idx = {_key(r): r for r in rs}
    n_add = n_upd = 0
    for g in gold:
        k = _key(g)
        if k in idx:
            row = idx[k]
            changed = any(row.get(c) != g.get(c) for c in COLS)
            if changed:
                row.update({c: g.get(c) for c in COLS})
                n_upd += 1
        else:
            rs.append(g)
            idx[k] = g
            n_add += 1
    RS_JSON.write_text(json.dumps(rs, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"rate-sens gold applied: {n_add} added, {n_upd} updated ({len(rs)} rows)")
    return 0


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "apply"
    if mode == "build":
        xp = sys.argv[2] if len(sys.argv) > 2 else str(REPO / "insurequant_master_tables.xlsx")
        sys.exit(build(xp))
    sys.exit(apply())
