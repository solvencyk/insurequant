#!/usr/bin/env python3
"""Parse DB손해보험 (KR0011) IR factsheets into an IR-derived NB CSM multiple series.

DB does not publish a ready 'CSM 배수', but its factsheet gives the two components:
  - numerator : 'BEL,CSM변동' sheet -> CSM section -> '신계약 유입' (per single quarter, 백만원)
  - denominator: '신규월납' sheet -> monthly '월납신규' (월납초회) + '비월납' (기타초회) (억원)
We derive a YTD-cumulative multiple (same basis as the KIDI-computed series and the
삼성/메리츠/한화생명 disclosed series): multiple = cumYTD(신계약CSM) / cumYTD(월납신규+비월납).

Numerator cross-checks EXACTLY against the KIDI/DART numerator (e.g. 24.1Q = 7,175.2억),
and the denominator (월납신규+비월납) cross-checks closely against KIDI 월납초회+기타초회.

Output: data/ir/series/KR0011_DB손해보험.json
"""
import json
import glob
import os
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

KR = "KR0011"
NAME = "DB손해보험"
GLOB = r"data/ir/FY*/KR0011_DB손해보험/*actSheet*.xls*"  # FactSheet files only (skip IR Report PDFs)


def _q(year, qn):
    return f"{year}.{qn}Q"


def parse_csm(ws):
    """Return {'2024.1Q': csm_baekman, ...} from the (2) CSM section, 신계약 유입 row."""
    out = {}
    year = None
    in_csm = False
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        s1 = v1.strip() if isinstance(v1, str) else ""
        s2 = v2.strip() if isinstance(v2, str) else ""
        if "(2) CSM" in s1 or "(2) CSM" in s2:
            in_csm = True
        if not in_csm:
            continue
        # header row: '구분' | 'YY.1Q' | '2Q' | '3Q' | '4Q'
        if s2 == "구분":
            h = ws.cell(r, 3).value
            m = re.search(r"(\d{2})\.1Q", str(h)) if h is not None else None
            if m:
                year = 2000 + int(m.group(1))
        if "신계약 유입" in s2 and year is not None:
            for i, qn in enumerate(("1", "2", "3", "4")):
                val = ws.cell(r, 3 + i).value
                if isinstance(val, (int, float)) and val != 0:
                    out[_q(year, qn)] = float(val)  # 백만원
            break
    return out


def parse_premium(ws):
    """Return {year:int -> {'wolnap':[12 months 억], 'etc':[12 months 억]}} from 신규월납."""
    blocks = {}
    cur_year = None
    for r in range(1, ws.max_row + 1):
        v2 = ws.cell(r, 2).value
        v3 = ws.cell(r, 3).value
        # year header row
        if isinstance(v2, (int, float)) and 2020 <= int(v2) <= 2030:
            cur_year = int(v2)
            blocks.setdefault(cur_year, {})
            continue
        if isinstance(v2, str) and re.fullmatch(r"20\d{2}", v2.strip()):
            cur_year = int(v2.strip())
            blocks.setdefault(cur_year, {})
            continue
        if cur_year is None:
            continue
        label2 = v2.strip() if isinstance(v2, str) else ""
        label3 = v3.strip() if isinstance(v3, str) else ""
        months = [ws.cell(r, 4 + i).value for i in range(12)]  # col4..col15 = 1월..12월
        months = [float(x) if isinstance(x, (int, float)) else 0.0 for x in months]
        if label2 == "월납신규":
            blocks[cur_year]["wolnap"] = months
        elif label3 == "비월납":
            blocks[cur_year]["etc"] = months
    return blocks


def qsum(months, qn):
    """Sum a quarter's 3 months (1-indexed quarter)."""
    i0 = (qn - 1) * 3
    return sum(months[i0:i0 + 3])


def main():
    files = sorted(glob.glob(GLOB))
    csm = {}          # 'YYYY.nQ' -> 백만원 (single quarter)
    prem = {}         # year -> {'wolnap':[...], 'etc':[...]}
    used = []
    for f in files:
        if not f.lower().endswith(".xlsx"):
            continue  # .xls (old format) — later .xlsx files carry the same history
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            print(f"  skip {os.path.basename(f)}: {type(e).__name__}")
            continue
        if "BEL,CSM변동" in wb.sheetnames:
            for k, v in parse_csm(wb["BEL,CSM변동"]).items():
                csm[k] = v  # later (sorted) files override -> revised values win
        if "신규월납" in wb.sheetnames:
            for y, d in parse_premium(wb["신규월납"]).items():
                # merge: keep the most complete (max non-zero months)
                cur = prem.get(y, {})
                for key in ("wolnap", "etc"):
                    if key in d:
                        if key not in cur or sum(1 for x in d[key] if x) >= sum(1 for x in cur[key] if x):
                            cur[key] = d[key]
                prem[y] = cur
        used.append(os.path.basename(f))

    # build per-quarter single-Q components, then YTD cumulative
    series = {}
    for y in sorted({int(k[:4]) for k in csm}):
        cum_num = 0.0
        cum_den = 0.0
        for qn in (1, 2, 3, 4):
            key = _q(y, qn)
            if key not in csm:
                continue
            num_q = csm[key] / 100.0  # 백만원 -> 억원
            pb = prem.get(y, {})
            wol = qsum(pb.get("wolnap", [0] * 12), qn)
            etc = qsum(pb.get("etc", [0] * 12), qn)
            den_q = wol + etc  # 억원 (월납초회 + 기타초회)
            cum_num += num_q
            cum_den += den_q
            series[key] = {
                "multiple_derived": round(cum_num / cum_den, 2) if cum_den else None,
                "nb_csm_eok": round(cum_num, 1),
                "premium_eok": round(cum_den, 1),
                "premium_wolnap_eok": round(wol, 1),    # this quarter only (월납초회)
                "premium_etc_eok": round(etc, 1),       # this quarter only (기타초회=비월납)
                "basis": "derived: cumYTD(신계약CSM 유입) ÷ cumYTD(월납신규+비월납)",
            }

    payload = {
        "company": NAME,
        "kr": KR,
        "sector": "Non-Life",
        "metric": "IR-derived NB CSM multiple (배) = 누계 신계약 CSM 유입 ÷ 누계(월납초회+기타초회). "
                  "DB IR factsheet 'BEL,CSM변동'(신계약 유입) + '신규월납'(월납신규+비월납).",
        "units": {"nb_csm_eok": "억원 (YTD)", "premium_eok": "억원 (YTD)", "multiple_derived": "배"},
        "source_files": used,
        "series": dict(sorted(series.items())),
    }
    out = Path("data/ir/series") / f"{KR}_{NAME}.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {out}  ({len(series)} quarters from {len(used)} files)")
    for k, v in payload["series"].items():
        print(f"  {k}: mult={v['multiple_derived']}  num={v['nb_csm_eok']}  den={v['premium_eok']} "
              f"(월납{v['premium_wolnap_eok']}+기타{v['premium_etc_eok']})")


if __name__ == "__main__":
    main()
