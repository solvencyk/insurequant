# -*- coding: utf-8 -*-
"""Cherry-pick sync of ONE sheet in insurequant_master_tables.xlsx from its master JSON.

owner 지시 (inbox `_resolved/20260819T0500Z` L119, 2026-08-19):
    "`insurequant_master_tables.xlsx`는 **전체 재생성 금지** — <해당> 시트만 cherry-pick 동기화할 것."
    "니가 뽑은 행이 RED=0이면 그거 기준으로 통일. 그래도 기존이랑 비교해서 진짜 맞는 행만 덮어쓰기로 동기화해"

`build_master_xlsx.py` 는 `ExcelWriter(mode="w")` 로 **파일 전체를 새로 쓴다** — MASTERS 목록 밖의
수기 시트(피벗 등)를 지우고, 다른 레인이 손으로 맞춰 둔 시트·설명까지 되돌린다. 그래서 마스터 JSON
하나가 바뀌었을 때는 이 스크립트로 **그 시트의 바뀐 셀·추가된 행·삭제된 행만** 반영한다.

방식:
  1. 대상 시트의 목표 상태를 `build_master_xlsx` 의 `MASTERS`/`coerce` 를 **import 해서** 만든다
     (여기서 스키마를 다시 구현하면 빌더가 바뀌는 순간 어긋난다).
  2. 행 식별키 = 값이 아닌 컬럼 전부(`TEXT_COLS` + `항목번호`). 그 키 시퀀스를 `difflib` 로 정렬해
     equal / insert / delete 로 나눈다.
  3. equal 구간은 **값 컬럼만 셀 단위로 비교**해 다른 셀만 덮어쓴다. insert 는 그 자리에 행을 넣고
     이웃 행의 서식(글꼴·표시형식)을 복사한다. delete 는 행을 지운다.
  4. 사후 검증 — (a) 동기화한 시트가 목표와 **셀 단위로 완전 일치**, (b) **나머지 시트는 값 기준
     완전 동일**(동기화 전 스냅샷과 대조). 하나라도 어긋나면 저장하지 않고 실패로 끝낸다.
  5. `요약` 시트의 행수는 **실제 시트 행수로** 맞춘다(설명 칸은 손대지 않는다 — 다른 레인이 손으로
     고쳐 둔 문구가 있다).

전제: 이 워크북에는 수식이 0개다(확인 필수 — 수식이 있으면 openpyxl 재저장이 캐시값을 날린다).
수식이 하나라도 발견되면 실행을 거부한다. 피벗/차트/도형/테이블도 없어야 한다(openpyxl 이 떨군다).

Usage:
  C:/Users/sangwook.cho/venvs/insurequant/Scripts/python.exe \
      scripts/sync_master_xlsx_sheet.py "K-ICS공시" [--dry-run]
"""
from __future__ import annotations

import io
import json
import sys
import zipfile
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "scripts"))

from build_master_xlsx import FLATTEN, FONT, MASTERS, NUMERIC_COLS, TEXT_COLS, coerce  # noqa: E402

XLSX = REPO / "insurequant_master_tables.xlsx"
SUMMARY_SHEET = "요약"
RISKY_PARTS = ("pivot", "chart", "drawing", "media", "/table", "vml", "queryTable", "connections")


def target_rows(json_file: str) -> tuple[list[str], list[list]]:
    """Build the sheet's target state exactly as build_master_xlsx would write it."""
    data = json.loads((REPO / json_file).read_text(encoding="utf-8"))
    if json_file in FLATTEN:  # tier1/tier2/forward_capital: reshape to long-format first
        data = FLATTEN[json_file](data)
    df = coerce(pd.DataFrame(data))
    cols = list(df.columns)
    rows = []
    for rec in df.to_dict("records"):
        out = []
        for c in cols:
            v = rec[c]
            if v is None or v is pd.NA or (isinstance(v, float) and pd.isna(v)):
                out.append(None)                      # pandas writes NA/NaN as an empty cell
            elif c in NUMERIC_COLS or c == "항목번호":
                f = float(v)
                out.append(int(f) if f.is_integer() else f)  # Excel stores whole floats as ints
            else:
                out.append(str(v))
        rows.append(out)
    return cols, rows


def sheet_rows(ws, ncol: int) -> list[list]:
    return [[ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
            for r in range(2, ws.max_row + 1)]


def norm(v):
    """Compare-normalize: Excel round-trips whole floats to int, and blanks to None.

    `coerce` 는 값이 아닌 열을 전부 문자열로 만든다(`레벨` -> '1'). 그런데 owner 가 워크북을
    Excel 로 열어 저장하면 숫자처럼 보이는 그 텍스트가 **숫자 1 로 바뀌어** 저장된다. 그러면
    시트(int 1) 와 목표(str '1') 의 키가 영원히 어긋나 difflib 이 정렬을 포기하고 17BS 시트
    전체를 delete+insert 하겠다고 보고했다(2026-08-21 실측: 삭제 6855 · 추가 6852). 앞뒤로
    한 번 왕복해 같은 표기로 돌아오는 것만 숫자로 본다 — 티커 '000060' 은 int 로 가면 '60' 이
    되어 왕복이 깨지므로 문자열로 남는다(선행 0 보존).
    """
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str) and v.lstrip("-").isdigit() and str(int(v)) == v:
        return int(v)
    if isinstance(v, float):
        # xlsx 는 float 를 유효숫자 15자리로 쓴다(openpyxl `%.15g`). 마스터 JSON 의
        # 168.79000000000002 는 시트에 168.79 로 밖에 못 들어가므로, 16자리째만 다른 것을
        # '차이'로 세면 **영원히 수렴하지 않는 EDIT 34칸**이 매번 보고된다(2026-08-21 실측:
        # 동기화 직후 재확인에서 K-ICS 1 · 금리민감도 7 · 손익분해PL 26). 시트가 담을 수 있는
        # 정밀도로 맞춰 비교한다 — 진짜 값 차이가 16자리째에서만 나는 일은 없다.
        return float("%.15g" % v)
    return v


def key_of(row: list, cols: list[str], key_idx: list[int]) -> tuple:
    return tuple(norm(row[i]) for i in key_idx)


def preflight() -> None:
    with zipfile.ZipFile(XLSX) as z:
        bad = [n for n in z.namelist() if any(k in n for k in RISKY_PARTS)]
    if bad:
        sys.exit(f"REFUSE: 워크북에 openpyxl이 떨구는 파트가 있다 {bad} — 손으로 처리할 것")


def main() -> int:
    # 2026-09-02: 이 줄은 모듈 최상위에 있었다. `scripts/check_master_xlsx_drift.py` 가
    # `target_rows`/`norm`/`key_of` 를 **import 해서** 쓰는데(스키마 재타이핑 금지), import
    # 만으로 호출자의 stdout 이 통째로 바뀌면 안 된다 — pytest 캡처·게이트 출력에 부작용이
    # 생긴다. CLI 로 돌 때만 적용되도록 main() 안으로 옮겼다(동작 동일).
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    dry = "--dry-run" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if len(args) != 1:
        sys.exit(__doc__)
    sheet = args[0]

    spec = next((m for m in MASTERS if m[1] == sheet), None)
    if spec is None:
        sys.exit(f"REFUSE: '{sheet}' 는 build_master_xlsx.MASTERS 에 없다 "
                 f"(있는 것: {[m[1] for m in MASTERS]})")
    json_file = spec[0]

    preflight()
    wb = load_workbook(XLSX, data_only=False)

    n_formula = sum(
        1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    )
    if n_formula:
        sys.exit(f"REFUSE: 수식 {n_formula}개 발견 — openpyxl 재저장은 캐시값을 날린다")

    tgt_cols, tgt = target_rows(json_file)

    # 2026-08-29 (inbox/parser/20260829T0100Z): sync was built to cherry-pick an
    # EXISTING sheet — it has no path for a sheet that isn't in the workbook yet
    # (`wb[sheet]` would KeyError). Extend it minimally: create the sheet + header row
    # now (in-memory only — nothing hits disk until wb.save() below, which the existing
    # `if dry: return 0` further down already skips) so every downstream line
    # (cols/cur/diff/insert) sees the same "existing sheet, 0 data rows" shape it
    # already knows how to handle — SequenceMatcher(a=[], b=tgt_keys) naturally becomes
    # one big insert block, same code path as adding rows to a populated sheet.
    is_new_sheet = sheet not in wb.sheetnames
    if is_new_sheet:
        ws = wb.create_sheet(sheet)
        ws.append(tgt_cols)
    else:
        ws = wb[sheet]

    cols = [c.value for c in ws[1]]
    if cols != tgt_cols:
        sys.exit(f"REFUSE: 컬럼 불일치\n  시트: {cols}\n  마스터: {tgt_cols}")

    ncol = len(cols)
    key_idx = [i for i, c in enumerate(cols) if c in TEXT_COLS or c == "항목번호"]
    val_idx = [i for i in range(ncol) if i not in key_idx]
    print(f"sheet={sheet!r}  master={json_file}")
    print(f"  키 컬럼: {[cols[i] for i in key_idx]}")
    print(f"  값 컬럼: {[cols[i] for i in val_idx]}")

    cur = sheet_rows(ws, ncol)
    others_before = {w.title: sheet_rows(w, w.max_column) for w in wb.worksheets if w.title != sheet}

    cur_keys = [key_of(r, cols, key_idx) for r in cur]
    tgt_keys = [key_of(r, cols, key_idx) for r in tgt]
    sm = SequenceMatcher(a=cur_keys, b=tgt_keys, autojunk=False)

    cell_edits, inserts, deletes = [], [], []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for off in range(i2 - i1):
                for k in val_idx:
                    if norm(cur[i1 + off][k]) != norm(tgt[j1 + off][k]):
                        cell_edits.append((i1 + off, k, cur[i1 + off][k], tgt[j1 + off][k]))
        elif tag == "insert":
            inserts.extend((i1, tgt_keys[j], tgt[j]) for j in range(j1, j2))
        elif tag == "delete":
            deletes.extend((i, cur_keys[i]) for i in range(i1, i2))
        else:  # replace — identity changed; surface it rather than silently rewriting
            deletes.extend((i, cur_keys[i]) for i in range(i1, i2))
            inserts.extend((i2, tgt_keys[j], tgt[j]) for j in range(j1, j2))

    print(f"  변경 셀 {len(cell_edits)} · 추가 행 {len(inserts)} · 삭제 행 {len(deletes)} "
          f"(시트 {len(cur)}행 → 목표 {len(tgt)}행)")
    for i, k, old, new in cell_edits[:40]:
        idk = tuple(cur[i][x] for x in key_idx)
        print(f"    EDIT   row{i + 2} {cols[k]}: {old!r} -> {new!r}   {idk}")
    if len(cell_edits) > 40:
        print(f"    ... +{len(cell_edits) - 40} more")
    for pos, k, _row in inserts:
        print(f"    INSERT @row{pos + 2} {k}")
    for i, k in deletes:
        print(f"    DELETE  row{i + 2} {k}")

    if dry:
        print("(dry-run; 파일 안 씀)")
        return 0
    if not (cell_edits or inserts or deletes):
        print("이미 동기 상태 — 파일 안 씀")
        return 0

    # --- apply: cell edits first (row indices still valid), then deletes desc, then inserts asc
    for i, k, _old, new in cell_edits:
        ws.cell(row=i + 2, column=k + 1).value = new

    for i, _k in sorted(deletes, key=lambda t: -t[0]):
        ws.delete_rows(i + 2, 1)
    shift = {i for i, _k in deletes}

    applied = 0        # 이미 끼워 넣은 행 수 — 안 세면 두 번째 이후 삽입이 그만큼 앞으로 밀린다
    for pos, _k, row in sorted(inserts, key=lambda t: t[0]):
        at = pos - sum(1 for d in shift if d < pos) + applied + 2
        applied += 1
        ws.insert_rows(at, 1)
        donor = at + 1 if at + 1 <= ws.max_row else at - 1
        for c in range(1, ncol + 1):
            cell = ws.cell(row=at, column=c)
            cell.value = row[c - 1]
            src = ws.cell(row=donor, column=c)
            cell.font = copy(src.font)
            cell.alignment = copy(src.alignment)
            cell.border = copy(src.border)
            cell.number_format = src.number_format

    nrow = ws.max_row
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ncol).column_letter}{nrow}"

    if is_new_sheet:
        # The insert loop above copies neighbor formatting per row (correct for
        # cherry-picking a FEW rows into an already-styled sheet); on a brand-new sheet
        # there is no real neighbor yet, so every row ends up copying the header's
        # bold-white-on-blue look. Fix it up once here — same header/body/column-width/
        # freeze-pane styling build_master_xlsx.py applies (kept in sync by eye; that
        # script is never *run* per this ticket's constraints, so there is no import to
        # share instead).
        thin = Side(style="thin", color="D9D9D9")
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name=FONT, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        base_w = {"원수사명": 22, "항목명": 30, "비고": 60}
        for c, name in enumerate(cols, start=1):
            letter = ws.cell(row=1, column=c).column_letter
            for r in range(2, nrow + 1):
                cl = ws.cell(row=r, column=c)
                cl.font = Font(name=FONT)
                cl.alignment = Alignment()
                cl.border = Border()
                if name in NUMERIC_COLS and isinstance(cl.value, (int, float)):
                    cl.number_format = "#,##0.##;(#,##0.##);-"
                else:
                    cl.number_format = "General"
            ws.column_dimensions[letter].width = base_w.get(name, max(11, min(28, len(str(name)) + 6)))
        ws.freeze_panes = "A2"

    # --- 요약: 신규 시트면 '합계' 행 바로 위에 행을 추가(설명은 MASTERS 의 desc 로 최초 1회만
    # 채운다 — 이후에는 기존 규칙대로 손대지 않는다), 그 다음 행수는 항상 실측으로 갱신
    # (설명 칸은 다른 레인이 손으로 고쳐 둬서 손대지 않는다)
    fixed = []
    added_summary_row = False
    if SUMMARY_SHEET in wb.sheetnames:
        idx = wb[SUMMARY_SHEET]
        listed = {idx.cell(row=r, column=1).value for r in range(4, idx.max_row + 1)}
        if sheet not in listed:
            total_row = next((r for r in range(4, idx.max_row + 1)
                               if idx.cell(row=r, column=1).value == "합계"), idx.max_row + 1)
            idx.insert_rows(total_row, 1)
            donor = total_row - 1
            values = [sheet, json_file, len(tgt), spec[2]]
            for c in range(1, 5):
                cell = idx.cell(row=total_row, column=c)
                cell.value = values[c - 1]
                src = idx.cell(row=donor, column=c)
                cell.font = copy(src.font)
                cell.alignment = copy(src.alignment)
                cell.number_format = src.number_format
            added_summary_row = True
        total = 0
        for r in range(4, idx.max_row + 1):
            name = idx.cell(row=r, column=1).value
            if name == "합계":
                idx.cell(row=r, column=3).value = total
                continue
            if name in wb.sheetnames:
                actual = wb[name].max_row - 1
                if idx.cell(row=r, column=3).value != actual:
                    fixed.append((name, idx.cell(row=r, column=3).value, actual))
                idx.cell(row=r, column=3).value = actual
                total += actual

    # --- verify BEFORE saving
    after = sheet_rows(ws, ncol)
    bad = [(i, [(cols[k], a[k], b[k]) for k in range(ncol) if norm(a[k]) != norm(b[k])])
           for i, (a, b) in enumerate(zip(after, tgt)) if any(norm(a[k]) != norm(b[k]) for k in range(ncol))]
    if len(after) != len(tgt) or bad:
        sys.exit(f"ABORT(저장 안 함): 동기화 후 시트가 마스터와 불일치 — "
                 f"행수 {len(after)} vs {len(tgt)}, 불일치 행 {len(bad)}: {bad[:3]}")
    for w in wb.worksheets:
        if w.title == sheet or w.title == SUMMARY_SHEET:
            continue
        if [[norm(v) for v in r] for r in sheet_rows(w, w.max_column)] != \
           [[norm(v) for v in r] for r in others_before[w.title]]:
            sys.exit(f"ABORT(저장 안 함): 손대지 않아야 할 시트 {w.title!r} 가 변했다")

    wb.save(XLSX)
    print(f"  검증 OK — {sheet} {len(after)}행 × {ncol}열 마스터와 완전 일치, 나머지 시트 값 동일")
    if added_summary_row:
        print(f"  요약 시트에 신규 행 추가: {sheet}")
    if fixed:
        print("  요약 행수 실측 보정:")
        for name, old, new in fixed:
            print(f"    {name}: {old} -> {new}")
    print(f"wrote {XLSX.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
