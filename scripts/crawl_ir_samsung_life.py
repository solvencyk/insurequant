#!/usr/bin/env python3
"""삼성생명 IR: download ALL fact sheet (XLS) files via Selenium, parse sheet Ⅰ-5
(CSM 상세) → disclosed 신계약 CSM 배수 (월초대비) quarterly series, 2023.1Q+.

Output: data/ir/decks/samsung_life/series.json + raw .xlsx files.
The factsheet is the clean source (no PDF parsing). Each factsheet spans the
current + prior FY, so all of them merged cover 2023.1Q ~ latest.
"""
import json
import sys
import time
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

sys.stdout.reconfigure(encoding="utf-8")
URL = "https://www.samsunglife.com/individual/display/invest/PDK-IRIVI015220M"
DL = Path("data/ir/decks/samsung_life").resolve()
DL.mkdir(parents=True, exist_ok=True)


def download_all_factsheets() -> list[Path]:
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,2400")
    opts.add_experimental_option("prefs", {
        "download.default_directory": str(DL), "download.prompt_for_download": False,
        "download.directory_upgrade": True, "safebrowsing.enabled": True,
    })
    d = webdriver.Chrome(options=opts)
    saved = []
    try:
        d.get(URL)
        time.sleep(7)
        d.execute_cdp_cmd("Page.setDownloadBehavior",
                          {"behavior": "allow", "downloadPath": str(DL)})
        n = len(d.find_elements(By.XPATH, "//*[normalize-space(text())='XLS 다운로드']"))
        print(f"{n} XLS buttons")
        for i in range(n):
            btns = d.find_elements(By.XPATH, "//*[normalize-space(text())='XLS 다운로드']")
            if i >= len(btns):
                break
            before = {p.name for p in DL.glob('*.xlsx')}
            d.execute_script("arguments[0].scrollIntoView({block:'center'});", btns[i])
            d.execute_script("arguments[0].click();", btns[i])
            got = None
            for _ in range(40):
                time.sleep(1)
                if any(f.name.endswith('.crdownload') for f in DL.glob('*')):
                    continue
                new = [f for f in DL.glob('*.xlsx') if f.name not in before]
                if new:
                    got = new[0]
                    break
            if got:
                saved.append(got)
                print(f"  [{i}] {got.name}")
            else:
                print(f"  [{i}] download timeout")
    finally:
        d.quit()
    return saved


def parse_factsheet(path: Path) -> dict:
    """sheet Ⅰ-5 → {period 'YYYY.NQ': {nb_csm_eok, multiple_wolcho, multiple_ape,
    by_product:{사망,건강,연금저축}}}. Returns {} if sheet/labels not found."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for sh in wb.worksheets:
        if any(isinstance(c.value, str) and "월초대비 신계약CSM 배수" in c.value
               for row in sh.iter_rows() for c in row):
            ws = sh
            break
    if ws is None:
        return {}
    # locate header rows: FY labels row (contains 'FY25' etc.) + quarter row (1Q..)
    fy_row = q_row = None
    for r in range(1, 12):
        rowvals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        joined = " ".join(str(v) for v in rowvals if v)
        if "FY" in joined and fy_row is None and any(
                isinstance(v, str) and v.strip().startswith("FY") for v in rowvals):
            fy_row = r
        if q_row is None and sum(1 for v in rowvals if isinstance(v, str) and v.strip() in
                                 ("1Q", "2Q", "3Q", "4Q")) >= 2:
            q_row = r
    if not fy_row or not q_row:
        return {}
    # map column -> period, stopping at the '누적 기준' section
    cutoff = ws.max_column + 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(fy_row, c).value
        if isinstance(v, str) and "누적" in v:
            cutoff = c
            break
    col_fy = {}
    cur = None
    for c in range(1, cutoff):
        v = ws.cell(fy_row, c).value
        if isinstance(v, str) and v.strip().startswith("FY"):
            cur = v.strip()
        col_fy[c] = cur
    col_period = {}
    for c in range(1, cutoff):
        q = ws.cell(q_row, c).value
        fy = col_fy.get(c)
        if isinstance(q, str) and q.strip() in ("1Q", "2Q", "3Q", "4Q") and fy:
            yr = "20" + fy.strip()[2:4]
            col_period[c] = f"{yr}.{q.strip()}"

    def row_with(label_sub):
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip() == label_sub:
                    return c.row
        return None

    def subrow_after(start_row, label_sub):
        for r in range(start_row + 1, start_row + 6):
            v = ws.cell(r, 5).value
            if isinstance(v, str) and label_sub in v:
                return r
        return None

    r_csm = row_with("신계약 CSM")
    r_wol = row_with("월초대비 신계약CSM 배수 (배)")
    r_ape = row_with("APE대비 신계약CSM 배수 (%)")
    out = {}
    for c, per in col_period.items():
        rec = {}
        if r_csm and ws.cell(r_csm, c).value is not None:
            rec["nb_csm_eok"] = round(float(ws.cell(r_csm, c).value) * 10, 1)  # 십억→억
        if r_wol and ws.cell(r_wol, c).value is not None:
            rec["multiple_wolcho"] = round(float(ws.cell(r_wol, c).value), 2)
        if r_ape and ws.cell(r_ape, c).value is not None:
            rec["multiple_ape_pct"] = round(float(ws.cell(r_ape, c).value), 1)
        if rec:
            out[per] = rec
    return out


def main():
    files = download_all_factsheets() or sorted(DL.glob("SLI*Factsheet*.xlsx"))
    merged = {}
    for f in sorted(files):
        try:
            ser = parse_factsheet(f)
        except Exception as e:  # noqa: BLE001
            print(f"  parse fail {f.name}: {e}")
            continue
        for per, rec in ser.items():
            merged.setdefault(per, {}).update(rec)
            merged[per].setdefault("source_files", [])
            if f.name not in merged[per]["source_files"]:
                merged[per]["source_files"].append(f.name)
    payload = {
        "company": "삼성생명", "sector": "Life",
        "metric": "월초대비 신계약CSM 배수 (배) — 삼성생명 IR factsheet sheet Ⅰ-5 (별도)",
        "source": "samsunglife.com IR factsheet XLS",
        "series": dict(sorted(merged.items())),
    }
    out = DL / "series.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nwrote {out}")
    for per, rec in sorted(merged.items()):
        print(f"  {per}: 배수(월초)={rec.get('multiple_wolcho')}  신계약CSM={rec.get('nb_csm_eok')}억")


if __name__ == "__main__":
    main()
