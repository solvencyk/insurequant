#!/usr/bin/env python3
"""Parse 삼성화재해상보험 (KR0008) IR factsheets into a disclosed NB CSM multiple series.

Samsung Fire's factsheet 'CSM' sheet DISCLOSES the NB CSM multiple directly:
  - 'CSM' sheet -> '신계약 CSM' section -> 합계 row  : single-quarter 신계약 CSM (억원)
  - 'CSM' sheet -> '신계약 CSM 배수' section -> 합계 row: single-quarter disclosed 배수 (배)
  - 'Premiums & Persistency Ratio' -> '장기보험료 - 월납환산 신계약' -> 합계: monthly-avg 월납환산 (억원)
Columns E/F/G/H = Q1/Q2/Q3/Q4 of the file's FY (header row 'FY26' etc. gives the year);
column I = prior-year. Values are SINGLE-QUARTER (신계약 CSM 3Q>4Q proves non-cumulative);
the disclosed 배수 denominator = 월납환산(월평균) × 3 months  (6267.3 / 14.08 = 445.1 = 148.37×3).

We record both the disclosed single-Q 배수 and a YTD-cumulative derived multiple
(= cumYTD(신계약CSM) ÷ cumYTD(월납환산×3)) so it can be compared to the KIDI-computed series.
The numerator cross-checks EXACTLY vs KIDI/DART (2023 cumsum = 6782.7/14426/26067.9/34995.3).

Output: data/ir/series/KR0008_삼성화재해상보험.json
"""
import json
import glob
import os
import re
import sys
import zipfile
import tempfile
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

KR = "KR0008"
NAME = "삼성화재해상보험"
GLOB = r"data/ir/FY*/KR0008_삼성화재해상보험/*.xlsx"
COLS = {1: "E", 2: "F", 3: "G", 4: "H"}  # quarter -> column letter (E=5,F=6,G=7,H=8)


def safe_load(path):
    """openpyxl load with a workaround for Samsung Fire's malformed docProps/custom.xml."""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        tmp = os.path.join(tempfile.gettempdir(), "sf_" + os.path.basename(path))
        with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w") as zout:
            for it in zin.namelist():
                if it == "docProps/custom.xml":
                    continue
                data = zin.read(it)
                if it == "[Content_Types].xml":
                    data = re.sub(rb"<Override[^>]*custom\.xml[^>]*/>", b"", data)
                if it == "_rels/.rels":
                    data = re.sub(rb"<Relationship[^>]*custom\.xml[^>]*/>", b"", data)
                zout.writestr(it, data)
        return openpyxl.load_workbook(tmp, data_only=True)


def _c(ws, r, col):
    return ws.cell(r, col).value


def _txt(ws, r, col):
    v = ws.cell(r, col).value
    return v.strip() if isinstance(v, str) else ""


def parse_year(ws):
    """Find the current-year FY token in col E header area (e.g. 'FY26' -> 2026)."""
    for r in range(1, min(ws.max_row, 40) + 1):
        v = _c(ws, r, 5)  # col E
        if isinstance(v, str):
            m = re.search(r"FY(\d{2})", v)
            if m:
                return 2000 + int(m.group(1))
    return None


def parse_csm_sheet(ws):
    """Return (year, {q: csm_singleQ}, {q: disclosed_multiple_singleQ})."""
    year = parse_year(ws)
    section = None
    csm_row = mult_row = None
    for r in range(1, ws.max_row + 1):
        cC = _txt(ws, r, 3)
        if cC == "신계약 CSM":
            section = "csm"
        elif cC == "신계약 CSM 배수":
            section = "mult"
        elif cC in ("CSM 상각률", "CSM Movement") or cC.startswith("주)"):
            if section in ("csm", "mult"):
                section = None
        elif cC == "합계":
            if section == "csm" and csm_row is None:
                csm_row = r
            elif section == "mult" and mult_row is None:
                mult_row = r
    csm = {}
    mult = {}
    for q, col in COLS.items():
        ci = 4 + q  # E=5
        if csm_row:
            v = _c(ws, csm_row, ci)
            if isinstance(v, (int, float)) and v != 0:
                csm[q] = float(v)
        if mult_row:
            v = _c(ws, mult_row, ci)
            if isinstance(v, (int, float)) and v != 0:
                mult[q] = float(v)
    return year, csm, mult


def parse_premium_sheet(ws):
    """Return {q: 월납환산 monthly-avg (억)} from '장기보험료 - 월납환산 신계약' 합계 row."""
    section = None
    wol_row = None
    for r in range(1, ws.max_row + 1):
        cC = _txt(ws, r, 3)
        if "월납환산 신계약" in cC:
            section = "wol"
        elif cC in ("유지율 - 보장성보험", "보종별 보험료") or cC.startswith("주)"):
            if section == "wol":
                section = None
        elif cC == "합계" and section == "wol":
            wol_row = r
            break
    out = {}
    if wol_row:
        for q, col in COLS.items():
            v = _c(ws, wol_row, 4 + q)
            if isinstance(v, (int, float)) and v != 0:
                out[q] = float(v)
    return out


def main():
    files = sorted(glob.glob(GLOB))
    by_year = {}  # year -> {q: {'csm':, 'mult_disc':, 'wol_m':}}
    used = []
    for f in files:
        try:
            wb = safe_load(f)
        except Exception as e:
            print(f"  skip {os.path.basename(f)}: {type(e).__name__} {e}")
            continue
        if "CSM" not in wb.sheetnames:
            continue
        year, csm, mult = parse_csm_sheet(wb["CSM"])
        prem_sheet = next((s for s in wb.sheetnames if "Premium" in s or "월납" in s), None)
        wol = parse_premium_sheet(wb[prem_sheet]) if prem_sheet else {}
        if not year:
            continue
        yd = by_year.setdefault(year, {})
        for q in (1, 2, 3, 4):
            if q in csm or q in mult or q in wol:
                rec = yd.setdefault(q, {})
                if q in csm:
                    rec["csm"] = csm[q]
                if q in mult:
                    rec["mult_disc"] = mult[q]
                if q in wol:
                    rec["wol_m"] = wol[q]
        used.append(os.path.basename(f))

    series = {}
    for year in sorted(by_year):
        cum_csm = 0.0
        cum_den = 0.0
        for q in (1, 2, 3, 4):
            rec = by_year[year].get(q)
            if not rec or "csm" not in rec:
                continue
            csm_q = rec["csm"]
            wol_q = rec.get("wol_m")
            den_q = wol_q * 3 if wol_q else None  # monthly-avg -> quarterly 월납환산
            cum_csm += csm_q
            if den_q:
                cum_den += den_q
            key = f"{year}.{q}Q"
            series[key] = {
                "multiple_disclosed": round(rec["mult_disc"], 2) if "mult_disc" in rec else None,
                "multiple_derived_ytd": round(cum_csm / cum_den, 2) if cum_den else None,
                "nb_csm_eok": round(cum_csm, 1),            # YTD cumulative
                "nb_csm_singleQ_eok": round(csm_q, 1),
                "premium_eok": round(cum_den, 1) if cum_den else None,  # YTD 월납환산 (×3)
                "premium_wolhwansan_monthly_eok": round(wol_q, 1) if wol_q else None,
                "basis": "disclosed: single-Q 신계약CSM배수 (삼성화재 factsheet 'CSM'!신계약 CSM 배수 합계); "
                         "derived_ytd: cumYTD(신계약CSM) ÷ cumYTD(월납환산 월평균 ×3)",
            }

    payload = {
        "company": NAME,
        "kr": KR,
        "sector": "Non-Life",
        "metric": "Samsung Fire DISCLOSED NB CSM multiple (배). multiple_disclosed = single-quarter "
                  "신계약 CSM 배수 합계 (factsheet). multiple_derived_ytd = YTD cumulative for KIDI comparison.",
        "units": {"nb_csm_eok": "억원 (YTD)", "premium_eok": "억원 (YTD, 월납환산×3)", "multiple": "배"},
        "source_files": used,
        "series": dict(sorted(series.items())),
    }
    out = Path("data/ir/series") / f"{KR}_{NAME}.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {out}  ({len(series)} quarters from {len(used)} files)")
    for k, v in payload["series"].items():
        print(f"  {k}: disclosed={v['multiple_disclosed']}  derived_ytd={v['multiple_derived_ytd']}  "
              f"csmYTD={v['nb_csm_eok']} (Q={v['nb_csm_singleQ_eok']})  월납환산_m={v['premium_wolhwansan_monthly_eok']}")


if __name__ == "__main__":
    main()
